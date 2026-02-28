import openpyxl
import sys

wb = openpyxl.load_workbook('vendbees_app/ventory_sheet.xlsx')
ws = wb['Stocks']

print("=== STOCKS SHEET STRUCTURE ===")
print("\nColumn Headers:")
headers = [cell.value for cell in ws[1]]
for i, h in enumerate(headers, 1):
    print(f"{i}. {h}")

print("\n=== First 15 Rows ===")
for row_idx in range(1, min(16, ws.max_row+1)):
    row_data = [cell.value for cell in ws[row_idx]]
    print(f"Row {row_idx}: {row_data}")

print(f"\n=== Total Rows: {ws.max_row} ===")
print(f"=== Total Columns: {ws.max_column} ===")

# Get all sheet names
print(f"\n=== All Sheets: {wb.sheetnames} ===")
