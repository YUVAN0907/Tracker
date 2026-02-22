import os
import requests
from dotenv import load_dotenv
import openpyxl

load_dotenv()

token_url = f"https://login.microsoftonline.com/{os.environ['SP_TENANT_ID']}/oauth2/v2.0/token"
r = requests.post(token_url, data={
    'client_id': os.environ['SP_CLIENT_ID'],
    'client_secret': os.environ['SP_CLIENT_SECRET'],
    'grant_type': 'client_credentials',
    'scope': 'https://graph.microsoft.com/.default'
})
token = r.json()['access_token']

url = f"https://graph.microsoft.com/v1.0/sites/{os.environ['SP_SITE_ID']}/drives/{os.environ['SP_DRIVE_ID']}/items/{os.environ['SP_FILE_ID']}/content"
r = requests.get(url, headers={'Authorization': f'Bearer {token}'}, stream=True)
print('Status:', r.status_code)
print('Content-Length:', r.headers.get('Content-Length'))

with open('fresh_test.xlsx', 'wb') as f:
    for chunk in r.iter_content(chunk_size=8192):
        f.write(chunk)

size = os.path.getsize('fresh_test.xlsx')
print('Downloaded size:', size)

try:
    wb = openpyxl.load_workbook('fresh_test.xlsx')
    print('Sheets:', wb.sheetnames)
    ws = wb['OUR_PO']
    print('OUR_PO columns:', [cell.value for cell in ws[1]])
    print('OUR_PO rows:', ws.max_row)
    for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
        print(row)
except Exception as e:
    print('Error:', e)
