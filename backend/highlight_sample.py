import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import warnings
warnings.filterwarnings('ignore')

excel_path = r'C:\Users\Bharani\OneDrive\Desktop\Tracker\backend\ventory_sheet.xlsx'

# Load workbook for formatting
wb = load_workbook(excel_path)

# Yellow highlight for sample data
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_font = Font(color='FF0000', bold=True)

# 1. Current_Stock - ALL rows are sample data (it was empty before)
if 'Current_Stock' in wb.sheetnames:
    ws = wb['Current_Stock']
    # Add header for marker column if not exists
    if ws.cell(1, 4).value != 'Sample_Data':
        ws.cell(1, 4, 'Sample_Data')
        ws.cell(1, 4).font = Font(bold=True)
    
    # Mark all data rows as sample
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 4, 'SAMPLE')
        ws.cell(row, 4).fill = yellow_fill
        ws.cell(row, 4).font = red_font
        # Also highlight the entire row
        for col in range(1, 4):
            ws.cell(row, col).fill = yellow_fill
    
    print(f'Current_Stock: Marked {ws.max_row - 1} rows as SAMPLE (ALL are sample - sheet was empty)')

# 2. Sales_Log - Need to identify which are original vs sample
if 'Sales_Log' in wb.sheetnames:
    ws = wb['Sales_Log']
    # Original data was from 2026-01-07 and 2026-01-08 only (about 4 rows)
    # Sample data is from 2026-02-10 to 2026-02-16
    
    sample_count = 0
    original_count = 0
    
    # Add header for marker column
    if ws.cell(1, 7).value != 'Sample_Data':
        ws.cell(1, 7, 'Sample_Data')
        ws.cell(1, 7).font = Font(bold=True)
    
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row, 1).value  # Date is column A
        if date_val:
            date_str = str(date_val)[:10] if date_val else ''
            # Original dates were 2026-01-07 and 2026-01-08
            if date_str in ['2026-01-07', '2026-01-08']:
                ws.cell(row, 7, 'ORIGINAL')
                original_count += 1
            else:
                ws.cell(row, 7, 'SAMPLE')
                ws.cell(row, 7).fill = yellow_fill
                ws.cell(row, 7).font = red_font
                # Highlight the row
                for col in range(1, 7):
                    ws.cell(row, col).fill = yellow_fill
                sample_count += 1
    
    print(f'Sales_Log: {original_count} ORIGINAL rows, {sample_count} SAMPLE rows (highlighted)')

wb.save(excel_path)
print(f'File saved: {excel_path}')
print('Yellow highlighted rows = Sample data added by assistant')
