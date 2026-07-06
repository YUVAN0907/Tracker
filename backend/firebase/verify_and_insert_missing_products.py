#!/usr/bin/env python3
"""
Script to identify and insert missing products from Excel into Firebase
"""

import openpyxl
import sys
import re
from dataconnect_db import execute_graphql

def parse_self_life(value):
    """Extract numeric value from self life field like '3 MONTH'"""
    if not value:
        return 0
    val_str = str(value).strip().upper()
    match = re.search(r'\d+', val_str)
    return int(match.group()) if match else 0

def parse_units(value):
    """Parse units field - handle '-', decimals, and strings"""
    if not value or str(value).strip() == '-':
        return 0
    try:
        return int(float(str(value).strip()))
    except:
        return 0

# Load Excel file
excel_file = r"C:\Users\Bharani\OneDrive\Documents\Tracker1\Inventory_sheet.xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=True)
product_sheet = wb['Product_Master']

print("=" * 80)
print("FINDING MISSING PRODUCTS")
print("=" * 80)

# Get all products from Excel
# Column order: CATEGORY, PRODUCT_ID, PRODUCT_NAME, PO, MRP, QUANTITY, UNITS, GST, VENDOR_ID, EAN NO, SELF LIFE
excel_products = {}
for row_idx, row in enumerate(product_sheet.iter_rows(min_row=2, max_row=product_sheet.max_row, values_only=True), start=2):
    if not row[1]:  # Skip if PRODUCT_ID (column 1) is empty
        break
    
    category = str(row[0]).strip() if row[0] else ""
    product_id = str(row[1]).strip() if row[1] else ""
    product_name = str(row[2]).strip() if row[2] else ""
    po = row[3]
    mrp = row[4]
    quantity = str(row[5]).strip() if row[5] else ""
    units = parse_units(row[6])
    gst = row[7]
    vendor_id = str(row[8]).strip() if row[8] else ""
    ean_no = str(row[9]).strip() if row[9] else ""
    self_life = parse_self_life(row[10])
    
    if product_id:
        excel_products[product_id] = {
            'productId': product_id,
            'productName': product_name,
            'category': category,
            'vendorId': vendor_id,
            'mrp': float(mrp) if mrp else 0,
            'quantity': quantity,
            'units': units,
            'gst': float(gst) if gst else 0,
            'eanNo': ean_no,
            'selfLife': self_life,
            'row': row_idx
        }

print(f"\n[OK] Found {len(excel_products)} products in Excel Product_Master sheet")

# Get all products from Firebase
firebase_products = {}
firebase_vendors = {}
try:
    # Fetch vendors first
    print("\nFetching vendors from Firebase...")
    for offset in range(0, 100, 100):
        query = f"""
        query {{
            vendors(limit: 100, offset: {offset}) {{
                vendorId
            }}
        }}
        """
        result = execute_graphql(query)
        vendors = result.get('vendors', []) or []
        if not vendors:
            break
        
        for vendor in vendors:
            firebase_vendors[vendor['vendorId']] = True
    
    print(f"[OK] Found {len(firebase_vendors)} vendors in Firebase")
    
    # Fetch with pagination
    for offset in range(0, 500, 100):
        query = f"""
        query {{
            products(limit: 100, offset: {offset}) {{
                productId
                productName
                category
                vendorId
                mrp
                quantity
                units
                gst
                eanNo
                selfLife
            }}
        }}
        """
        result = execute_graphql(query)
        products = result.get('products', []) or []
        if not products:
            break
        
        for product in products:
            firebase_products[product['productId']] = product
except Exception as e:
    print(f"Error fetching products from Firebase: {e}", file=sys.stderr)
    sys.exit(1)

print(f"\n[OK] Found {len(firebase_products)} products in Firebase")

# Find missing products
missing_products = []
for product_id, excel_data in excel_products.items():
    if product_id not in firebase_products:
        missing_products.append(excel_data)

print(f"\n[WARN] Found {len(missing_products)} missing products that need to be inserted")

if not missing_products:
    print("[OK] All products are already in Firebase!")
    sys.exit(0)

print("\n" + "=" * 80)
print("MISSING PRODUCTS DETAILS")
print("=" * 80)

for idx, product in enumerate(missing_products, 1):
    vendor_id = product['vendorId'] if product['vendorId'] else 'UNKNOWN'
    print(f"\n{idx}. {product['productId']} - {product['productName']}")
    print(f"   Category: {product['category']}")
    print(f"   Vendor ID: {vendor_id}")
    print(f"   MRP: {product['mrp']}")
    print(f"   Quantity: {product['quantity']}")
    print(f"   Units: {product['units']}")

print("\n" + "=" * 80)
print("INSERTING MISSING PRODUCTS")
print("=" * 80)

success_count = 0
error_count = 0
errors = []

for idx, product in enumerate(missing_products, 1):
    try:
        # Use vendor ID if it exists in Firebase, otherwise use UNKNOWN
        vendor_id = product['vendorId'] if product['vendorId'] else 'UNKNOWN'
        if vendor_id != 'UNKNOWN' and vendor_id not in firebase_vendors:
            print(f"\n[WARN] [{idx}/{len(missing_products)}] Vendor {vendor_id} not found in Firebase, using UNKNOWN instead")
            vendor_id = 'UNKNOWN'
        
        # Escape strings for GraphQL
        product_name = product['productName'].replace('\\', '\\\\').replace('"', '\\"')
        category = product['category'].replace('\\', '\\\\').replace('"', '\\"')
        quantity = product['quantity'].replace('\\', '\\\\').replace('"', '\\"')
        ean_no = product['eanNo'].replace('\\', '\\\\').replace('"', '\\"')
        
        mutation = f"""
        mutation {{
            product_insert(data: {{
                productId: "{product['productId']}"
                productName: "{product_name}"
                category: "{category}"
                vendorId: "{vendor_id}"
                mrp: {product['mrp']}
                quantity: "{quantity}"
                units: {product['units']}
                gst: {product['gst']}
                eanNo: "{ean_no}"
                selfLife: {product['selfLife']}
            }})
        }}
        """
        
        result = execute_graphql(mutation)
        
        print(f"\n[OK] [{idx}/{len(missing_products)}] Inserted: {product['productId']} - {product['productName']}")
        success_count += 1
        
    except Exception as e:
        error_msg = str(e)
        print(f"\n[ERROR] [{idx}/{len(missing_products)}] Error inserting {product['productId']}: {error_msg}")
        error_count += 1
        errors.append({
            'product_id': product['productId'],
            'error': error_msg
        })

print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"\nTotal Missing Products: {len(missing_products)}")
print(f"Successfully Inserted: {success_count}")
print(f"Failed Insertions: {error_count}")

if errors:
    print("\n[ERROR] Errors encountered:")
    for error in errors:
        print(f"  - {error['product_id']}: {error['error']}")

if error_count == 0:
    print("\n[OK] All missing products inserted successfully!")

# Final verification
print("\n" + "=" * 80)
print("FINAL VERIFICATION")
print("=" * 80)

try:
    final_count = 0
    for offset in range(0, 500, 100):
        query = f"""
        query {{
            products(limit: 100, offset: {offset}) {{
                productId
            }}
        }}
        """
        result = execute_graphql(query)
        products = result.get('products', []) or []
        if not products:
            break
        final_count += len(products)
    
    print(f"\n[OK] Final product count in Firebase: {final_count}")
    print(f"[OK] Expected product count from Excel: {len(excel_products)}")
    
    if final_count == len(excel_products):
        print(f"\n[SUCCESS] All {final_count} products are now in Firebase!")
    else:
        diff = len(excel_products) - final_count
        print(f"\n[WARN] Mismatch: {diff} products still missing")
        
except Exception as e:
    print(f"Error verifying final count: {e}")

print("\n" + "=" * 80)
