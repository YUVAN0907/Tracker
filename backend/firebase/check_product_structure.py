import openpyxl

excel_file = r"C:\Users\Bharani\OneDrive\Documents\Tracker1\Inventory_sheet.xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=True)
product_sheet = wb['Product_Master']

print("First 10 rows of Product_Master sheet:")
print("=" * 150)

for row_idx, row in enumerate(product_sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
    print(f"Row {row_idx}: {row}")

print("\n" + "=" * 150)
print(f"Total rows: {product_sheet.max_row}")
print(f"Total columns: {product_sheet.max_column}")

# Count non-empty rows
count = 0
for row in product_sheet.iter_rows(min_row=2, max_row=product_sheet.max_row, values_only=True):
    if row[0]:  # If first column is not empty
        count += 1

print(f"Total product rows (excluding header): {count}")
