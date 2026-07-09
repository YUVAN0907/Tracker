#!/usr/bin/env python3
"""
Script to:
1. Update unitCost in products table from PO column in product_master
2. Handle selfLife ranges (e.g., "9-12 MONTHS" -> 9)
3. Verify all product details match between Excel and Firebase
"""

import openpyxl
import sys
import re
from dataconnect_db import execute_graphql

def parse_self_life(value):
    """Extract lowest numeric value from self life field like '9-12 MONTHS' or '3 MONTH'"""
    if not value:
        return 0
    val_str = str(value).strip().upper()
    # Find all numbers in the string
    matches = re.findall(r'\d+', val_str)
    if matches:
        # Return the first (lowest) number in case of ranges like "9-12"
        return int(matches[0])
    return 0

def parse_units(value):
    """Parse units field - handle '-', decimals, and strings"""
    if not value or str(value).strip() == '-':
        return 0
    try:
        return int(float(str(value).strip()))
    except:
        return 0

def parse_float(value):
    """Parse float field - handle '-', decimals, and strings"""
    if not value or str(value).strip() == '-':
        return 0.0
    try:
        return float(str(value).strip())
    except:
        return 0.0

# Load Excel file
excel_file = r"C:\Users\Bharani\OneDrive\Documents\Tracker1\Inventory_sheet.xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=True)
product_sheet = wb['Product_Master']

print("=" * 100)
print("UPDATING UNIT COST & VERIFYING PRODUCT DETAILS")
print("=" * 100)

# Get all products from Excel
# Column order: CATEGORY, PRODUCT_ID, PRODUCT_NAME, PO, MRP, QUANTITY, UNITS, GST, VENDOR_ID, EAN NO, SELF LIFE
excel_products = {}
for row_idx, row in enumerate(product_sheet.iter_rows(min_row=2, max_row=product_sheet.max_row, values_only=True), start=2):
    if not row[1]:  # Skip if PRODUCT_ID (column 1) is empty
        break
    
    category = str(row[0]).strip() if row[0] else ""
    product_id = str(row[1]).strip() if row[1] else ""
    product_name = str(row[2]).strip() if row[2] else ""
    po = row[3]  # This is unitCost
    mrp = row[4]
    quantity = str(row[5]).strip() if row[5] else ""
    units = parse_units(row[6])
    gst = row[7]
    vendor_id = str(row[8]).strip() if row[8] else ""
    ean_no = str(row[9]).strip() if row[9] else ""
    self_life = parse_self_life(row[10])
    
    if product_id:
        excel_products[product_id] = {
            'productId': product_id,
            'productName': product_name,
            'category': category,
            'vendorId': vendor_id,
            'mrp': parse_float(mrp),
            'quantity': quantity,
            'units': units,
            'gst': parse_float(gst),
            'eanNo': ean_no,
            'selfLife': self_life,
            'unitCost': parse_float(po),
            'row': row_idx
        }

print(f"\n[OK] Loaded {len(excel_products)} products from Excel")

# Get all products from Firebase
firebase_products = {}
try:
    for offset in range(0, 500, 100):
        query = f"""
        query {{
            products(limit: 100, offset: {offset}) {{
                productId
                productName
                category
                vendorId
                mrp
                quantity
                units
                gst
                eanNo
                selfLife
                unitCost
            }}
        }}
        """
        result = execute_graphql(query)
        products = result.get('products', []) or []
        if not products:
            break
        
        for product in products:
            firebase_products[product['productId']] = product
except Exception as e:
    print(f"Error fetching products from Firebase: {e}", file=sys.stderr)
    sys.exit(1)

print(f"[OK] Fetched {len(firebase_products)} products from Firebase")

# Products to update with unitCost
products_to_update = []
for product_id, excel_data in excel_products.items():
    if product_id in firebase_products:
        products_to_update.append(excel_data)

print(f"\n[OK] Found {len(products_to_update)} products to check/update")

# Update unitCost and selfLife in Firebase
print("\n" + "=" * 100)
print("UPDATING UNIT COST & SELF LIFE")
print("=" * 100)

success_count = 0
error_count = 0
errors = []
self_life_updates = []

for idx, product in enumerate(products_to_update, 1):
    try:
        # Get current Firebase product
        fb_product = firebase_products[product['productId']]
        
        # Check if selfLife needs update (convert ranges)
        firebase_self_life = fb_product.get('selfLife', 0)
        excel_self_life = product['selfLife']
        self_life_changed = firebase_self_life != excel_self_life
        
        # Escape strings for GraphQL
        product_name = product['productName'].replace('\\', '\\\\').replace('"', '\\"')
        category = product['category'].replace('\\', '\\\\').replace('"', '\\"')
        quantity = product['quantity'].replace('\\', '\\\\').replace('"', '\\"')
        ean_no = product['eanNo'].replace('\\', '\\\\').replace('"', '\\"')
        
        mutation = f"""
        mutation {{
            product_update(key: {{ productId: "{product['productId']}" }}, 
            data: {{ 
                productName: "{product_name}",
                category: "{category}",
                vendorId: "{product['vendorId']}",
                mrp: {product['mrp']},
                quantity: "{quantity}",
                units: {product['units']},
                gst: {product['gst']},
                eanNo: "{ean_no}",
                selfLife: {product['selfLife']},
                unitCost: {product['unitCost']}
            }})
        }}
        """
        
        result = execute_graphql(mutation)
        
        if self_life_changed:
            self_life_updates.append({
                'productId': product['productId'],
                'before': firebase_self_life,
                'after': excel_self_life
            })
        
        success_count += 1
        
        if idx % 50 == 0:
            print(f"[OK] Updated {idx}/{len(products_to_update)} products...")
        
    except Exception as e:
        error_msg = str(e)
        print(f"\n[ERROR] [{idx}/{len(products_to_update)}] Error updating {product['productId']}: {error_msg}")
        error_count += 1
        errors.append({
            'product_id': product['productId'],
            'error': error_msg
        })

print(f"\n[OK] Updated {success_count}/{len(products_to_update)} products")

if self_life_updates:
    print(f"\n[INFO] SelfLife corrected for {len(self_life_updates)} products:")
    for update in self_life_updates[:10]:  # Show first 10
        print(f"  - {update['productId']}: {update['before']} -> {update['after']}")
    if len(self_life_updates) > 10:
        print(f"  ... and {len(self_life_updates) - 10} more")

# Verify all product details match
print("\n" + "=" * 100)
print("VERIFICATION - COMPARING EXCEL VS FIREBASE")
print("=" * 100)

# Fetch fresh data from Firebase
firebase_verify = {}
try:
    for offset in range(0, 500, 100):
        query = f"""
        query {{
            products(limit: 100, offset: {offset}) {{
                productId
                productName
                category
                vendorId
                mrp
                quantity
                units
                gst
                eanNo
                selfLife
                unitCost
            }}
        }}
        """
        result = execute_graphql(query)
        products = result.get('products', []) or []
        if not products:
            break
        
        for product in products:
            firebase_verify[product['productId']] = product
except Exception as e:
    print(f"Error fetching verification data: {e}")

mismatches = []
perfect_matches = 0

for product_id, excel_data in excel_products.items():
    if product_id not in firebase_verify:
        print(f"[WARN] Product {product_id} not found in Firebase!")
        continue
    
    fb_data = firebase_verify[product_id]
    
    # Compare fields
    mismatch_fields = []
    
    if excel_data['productName'] != fb_data.get('productName'):
        mismatch_fields.append(f"productName: '{excel_data['productName']}' vs '{fb_data.get('productName')}'")
    
    if excel_data['category'] != fb_data.get('category'):
        mismatch_fields.append(f"category: '{excel_data['category']}' vs '{fb_data.get('category')}'")
    
    if excel_data['vendorId'] != fb_data.get('vendorId'):
        mismatch_fields.append(f"vendorId: '{excel_data['vendorId']}' vs '{fb_data.get('vendorId')}'")
    
    if round(excel_data['mrp'], 2) != round(float(fb_data.get('mrp') or 0), 2):
        mismatch_fields.append(f"mrp: {excel_data['mrp']} vs {fb_data.get('mrp')}")
    
    if excel_data['quantity'] != fb_data.get('quantity'):
        mismatch_fields.append(f"quantity: '{excel_data['quantity']}' vs '{fb_data.get('quantity')}'")
    
    if excel_data['units'] != fb_data.get('units'):
        mismatch_fields.append(f"units: {excel_data['units']} vs {fb_data.get('units')}")
    
    if round(excel_data['gst'], 2) != round(float(fb_data.get('gst') or 0), 2):
        mismatch_fields.append(f"gst: {excel_data['gst']} vs {fb_data.get('gst')}")
    
    if excel_data['eanNo'] != fb_data.get('eanNo'):
        mismatch_fields.append(f"eanNo: '{excel_data['eanNo']}' vs '{fb_data.get('eanNo')}'")
    
    if excel_data['selfLife'] != fb_data.get('selfLife'):
        mismatch_fields.append(f"selfLife: {excel_data['selfLife']} vs {fb_data.get('selfLife')}")
    
    if round(excel_data['unitCost'], 2) != round(float(fb_data.get('unitCost') or 0), 2):
        mismatch_fields.append(f"unitCost: {excel_data['unitCost']} vs {fb_data.get('unitCost')}")
    
    if mismatch_fields:
        mismatches.append({
            'productId': product_id,
            'productName': excel_data['productName'],
            'fields': mismatch_fields
        })
    else:
        perfect_matches += 1

print(f"\n[OK] Perfect matches: {perfect_matches}/{len(excel_products)}")
print(f"[WARN] Mismatches found: {len(mismatches)}/{len(excel_products)}")

if mismatches:
    print(f"\nShowing first 20 mismatches:")
    for idx, mismatch in enumerate(mismatches[:20], 1):
        print(f"\n{idx}. {mismatch['productId']} - {mismatch['productName']}")
        for field in mismatch['fields']:
            print(f"   - {field}")
    
    if len(mismatches) > 20:
        print(f"\n... and {len(mismatches) - 20} more products with mismatches")

print("\n" + "=" * 100)
print("SUMMARY")
print("=" * 100)
print(f"\nTotal products in Excel: {len(excel_products)}")
print(f"Total products in Firebase: {len(firebase_verify)}")
print(f"Products updated/verified: {success_count}")
print(f"Update errors: {error_count}")
print(f"SelfLife ranges corrected: {len(self_life_updates)}")
print(f"Perfect detail matches: {perfect_matches}/{len(excel_products)}")
print(f"Products with mismatches: {len(mismatches)}/{len(excel_products)}")

if perfect_matches == len(excel_products):
    print("\n[SUCCESS] All product details match perfectly between Excel and Firebase!")
else:
    percentage = (perfect_matches / len(excel_products)) * 100
    print(f"\n[INFO] {percentage:.1f}% of products match perfectly")

if error_count == 0:
    print("[OK] No errors during update/verification")

print("\n" + "=" * 100)
