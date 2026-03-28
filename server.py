import os
import time
import threading
import tempfile
import requests
import pandas as pd
import numpy as np
from flask import Flask, jsonify, request
from flask_cors import CORS
from datetime import datetime
from dotenv import load_dotenv

# --------------------------------------------------
# LOAD ENV VARIABLES
# --------------------------------------------------
load_dotenv()

TENANT_ID = os.environ["SP_TENANT_ID"]
CLIENT_ID = os.environ["SP_CLIENT_ID"]
CLIENT_SECRET = os.environ["SP_CLIENT_SECRET"]
SITE_ID = os.environ["SP_SITE_ID"]
DRIVE_ID = os.environ["SP_DRIVE_ID"]
FILE_ID = os.environ["SP_FILE_ID"]

PORT = 3001
# Use Windows temp folder to avoid OneDrive sync issues
TEMP_EXCEL = os.path.join(tempfile.gettempdir(), "vendbees_temp_update.xlsx")

# Add lock to prevent concurrent downloads
LOAD_LOCK = threading.Lock()

# --------------------------------------------------
# FLASK APP
# --------------------------------------------------
app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

# --------------------------------------------------
# SHEET MAP
# --------------------------------------------------
SHEET_MAP = {
    'Products': 'Product_Master',
    'Machines': 'Machine_Master',
    'Stock': 'Current_Stock',
    'Sales': 'Sales_Log',
    'Purchases': 'Vendor_Purchase',
    'Refills': 'Machine_Refill_Log',
    'Vendors': 'Vendor_Master',
    'Warehouse': 'Warehouse_Stock',
    'Purchased_Products': 'Purchased_Products',
    'Stocks': 'Stocks',
    'Stock_Products': 'Stock_Products',
    'Stock_Assignments': 'Stock_Assignments',
    'OUR_PO': 'OUR_PO'
}

db = {}
last_sync = 0

# --------------------------------------------------
# SHAREPOINT AUTH
# --------------------------------------------------
def get_access_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type": "client_credentials",
        "scope": "https://graph.microsoft.com/.default"
    }
    r = requests.post(url, data=data)
    r.raise_for_status()
    return r.json()["access_token"]

# --------------------------------------------------
# SHAREPOINT FILE OPS
# --------------------------------------------------
def download_excel():
    token = get_access_token()
    headers = {"Authorization": f"Bearer {token}"}
    url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/items/{FILE_ID}/content"
    
    # Remove old file if it exists to avoid partial file issues
    if os.path.exists(TEMP_EXCEL):
        try:
            os.remove(TEMP_EXCEL)
            time.sleep(0.2)  # Small delay after deletion
        except Exception as e:
            print(f"⚠️  Could not remove old file: {e}")
    
    # Use streaming to avoid file corruption on Windows
    r = requests.get(url, headers=headers, stream=True)
    r.raise_for_status()
    
    # Write using streaming chunks
    total_size = 0
    with open(TEMP_EXCEL, "wb") as f:
        for chunk in r.iter_content(chunk_size=8192):
            if chunk:  # Filter out keep-alive new chunks
                f.write(chunk)
                total_size += len(chunk)
        f.flush()  # Force flush to disk
        os.fsync(f.fileno())  # Ensure file is synced to disk
    
    # Verify file was written
    if total_size < 5000:
        raise Exception(f"Downloaded file too small: {total_size} bytes")
    
    # Verify ZIP header
    with open(TEMP_EXCEL, "rb") as f:
        header = f.read(4)
        print(f"   File header (hex): {header.hex()}")
        if header != b'PK\x03\x04':
            print(f"   ⚠️  WARNING: File doesn't start with ZIP signature, got {header}")
            # Try to read first 100 bytes to debug
            f.seek(0)
            first_bytes = f.read(100)
            print(f"   First 100 bytes: {first_bytes[:100]}")
    
    # Additional wait to ensure file is fully available
    time.sleep(0.5)
    
    print(f"✔ Downloaded Excel to {TEMP_EXCEL} ({total_size} bytes)")

def upload_excel(max_retries=3, retry_delay=2):
    """Upload Excel to SharePoint with retry logic for locked files"""
    
    # Verify file exists before attempting upload
    if not os.path.exists(TEMP_EXCEL):
        raise Exception(f"[CRITICAL] Temp Excel file does not exist: {TEMP_EXCEL}")
    
    file_size = os.path.getsize(TEMP_EXCEL)
    
    for attempt in range(max_retries):
        try:
            token = get_access_token()
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/octet-stream"
            }
            url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/items/{FILE_ID}/content"
            
            # Verify file still exists and is readable
            if not os.path.exists(TEMP_EXCEL):
                raise Exception(f"[CRITICAL] Temp file disappeared: {TEMP_EXCEL}")
            
            if not os.access(TEMP_EXCEL, os.R_OK):
                raise Exception(f"[CRITICAL] Temp file not readable: {TEMP_EXCEL}")
            
            with open(TEMP_EXCEL, "rb") as f:
                r = requests.put(url, headers=headers, data=f)
            
            if r.status_code == 423:
                # File is locked - retry after delay
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                    retry_delay *= 2  # Exponential backoff
                    continue
                else:
                    raise Exception("SharePoint file is locked. Please close the Excel file and try again.")
            
            r.raise_for_status()
            return True
        except FileNotFoundError as e:
            raise
        except requests.exceptions.HTTPError as e:
            if "423" in str(e) and attempt < max_retries - 1:
                print(f"⚠ File locked, retrying in {retry_delay}s... (attempt {attempt + 1}/{max_retries})")
                time.sleep(retry_delay)
                retry_delay *= 2
                continue
            raise
    return False

def save_all():
    """Save all dataframes to Excel file and upload to SharePoint"""
    try:
        # Get temp directory
        temp_dir = tempfile.gettempdir()
        
        # Ensure temp directory exists
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir, exist_ok=True)
        
        # Remove old file if it exists
        if os.path.exists(TEMP_EXCEL):
            try:
                os.remove(TEMP_EXCEL)
            except Exception as old_file_err:
                pass
        
        # Prepare all data first (outside the context manager)
        sheets_to_write = {}
        for key, sheet_name in SHEET_MAP.items():
            if key in db:
                df = db[key]
                if not df.empty:
                    # Make a copy and fix column names
                    df_copy = df.copy()
                    df_copy.columns = df_copy.columns.astype(str)
                    sheets_to_write[sheet_name] = df_copy
        
        # Now write all at once
        try:
            with pd.ExcelWriter(TEMP_EXCEL, engine="openpyxl") as writer:
                for sheet_name, df in sheets_to_write.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as write_error:
            import traceback
            traceback.print_exc()
            raise
        
        # Verify file was created
        if not os.path.exists(TEMP_EXCEL):
            raise Exception(f"[CRITICAL] Excel file was NOT created at {TEMP_EXCEL}")
        
        # Skip polling before upload
        skip_polling_for(15)
        
        # Upload to SharePoint
        try:
            upload_result = upload_excel()
            if not upload_result:
                raise Exception("upload_excel returned False")
        except Exception as upload_error:
            import traceback
            traceback.print_exc()
            raise
        
        print(f"[OK] save_all() completed successfully\n")
        return True
        
    except Exception as e:
        print(f"[ERROR] save_all() failed: {e}\n")
        import traceback
        traceback.print_exc()
        raise

# --------------------------------------------------
# HELPERS
# --------------------------------------------------
def clean_df(df):
    if df.empty:
        return df

    df.columns = df.columns.astype(str).str.strip()
    df = df.replace({np.nan: None, np.inf: None, -np.inf: None})

    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)

    return df.dropna(how="all")

def df_to_safe_dict(df):
    if df.empty:
        return []
    return df.replace({np.nan: None}).to_dict(orient="records")

# --------------------------------------------------
# LOAD DATA FROM SHAREPOINT
# --------------------------------------------------
def load_data():
    global db, last_sync
    
    # Acquire lock to prevent concurrent downloads
    with LOAD_LOCK:
        max_retries = 3
        retry_delays = [2, 3, 5]  # Increased delays for file lock timeout
        
        for attempt in range(max_retries):
            try:
                download_excel()
                
                # Wait longer to ensure file is fully available after download
                delay = retry_delays[attempt]
                print(f"⏳ Waiting {delay}s for file to be available...")
                time.sleep(delay)
                
                # Verify file exists and has reasonable size
                if not os.path.exists(TEMP_EXCEL):
                    raise Exception(f"Excel file not found at {TEMP_EXCEL}")
                
                file_size = os.path.getsize(TEMP_EXCEL)
                if file_size < 5000:
                    raise Exception(f"Excel file too small ({file_size} bytes), likely incomplete")
                
                print(f"📁 Opening Excel file ({file_size} bytes)...")
            
                # Copy file to a temporary location to avoid file locking issues
                import shutil
                temp_copy = os.path.join(tempfile.gettempdir(), f"vendbees_read_{attempt}.xlsx")
                try:
                    shutil.copy2(TEMP_EXCEL, temp_copy)
                    excel_to_read = temp_copy
                except Exception as e:
                    print(f"   Could not create temp copy: {e}, will try to read original...")
                    excel_to_read = TEMP_EXCEL
                
                # Try to open and read the Excel file
                try:
                    # Try using openpyxl directly first
                    print(f"   Attempting to read with openpyxl...")
                    from openpyxl import load_workbook
                    
                    wb = load_workbook(excel_to_read, data_only=True)
                    
                    # Convert workbook to dataframes
                    for key, sheet in SHEET_MAP.items():
                        sheet_name = sheet
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            # Convert worksheet to dataframe
                            df = pd.DataFrame([cell.value for cell in row] for row in ws.iter_rows(values_only=True))
                            if not df.empty:
                                # Use first row as header
                                df.columns = df.iloc[0]
                                df = df[1:]
                                df = clean_df(df)
                            
                            if key == "Products" and "PRODUCT_ID" in df.columns:
                                df = df[df["PRODUCT_ID"].notna()]

                            db[key] = df.reset_index(drop=True)
                        else:
                            db[key] = pd.DataFrame()
                    
                    wb.close()
                    
                except Exception as openpyxl_err:
                    print(f"   openpyxl failed: {openpyxl_err}")
                    print(f"   Trying pandas ExcelFile as fallback...")
                    
                    with pd.ExcelFile(excel_to_read, engine="openpyxl") as xls:
                        for key, sheet in SHEET_MAP.items():
                            if sheet in xls.sheet_names:
                                df = clean_df(pd.read_excel(xls, sheet))

                                if key == "Products" and "PRODUCT_ID" in df.columns:
                                    df = df[df["PRODUCT_ID"].notna()]

                                db[key] = df.reset_index(drop=True)
                            else:
                                db[key] = pd.DataFrame()

                    if "Sales" in db and "Date" in db["Sales"].columns:
                        db["Sales"]["Date"] = pd.to_datetime(
                            db["Sales"]["Date"], errors="coerce"
                        ).dt.strftime("%Y-%m-%d")

                    # Ensure OUR_PO has correct columns (add missing ones in memory only - don't save on load)
                    if "OUR_PO" in db:
                        our_po = db["OUR_PO"]
                        expected_columns = [
                            "PO_ID", "Vendor_ID", "Created_Date", "Total_Amount", "Product_ID", "Product_Name", 
                            "No_of_Cases", "Units_Per_Case", "PO_Price", "Line_Total", "Status"
                        ]
                        
                        # Only add missing columns in memory (no save on load to avoid corruption)
                        if our_po.empty:
                            our_po = pd.DataFrame(columns=expected_columns)
                        else:
                            for col in expected_columns:
                                if col not in our_po.columns:
                                    if col == "Status":
                                        our_po[col] = "Pending"
                                    elif col == "Line_Total":
                                        our_po[col] = our_po.apply(
                                            lambda r: int(r.get("No_of_Cases", 0) or 0) * int(r.get("Units_Per_Case", 1) or 1), 
                                            axis=1
                                        )
                                    else:
                                        our_po[col] = None  # Use None instead of empty string
                            
                            # Ensure column order (in memory)
                            our_po = our_po[[col for col in expected_columns if col in our_po.columns]]
                        
                        db["OUR_PO"] = our_po
                    
                    # Log Stocks sheet info
                    if "Stocks" in db:
                        stocks = db["Stocks"]
                        print(f"✔ Stocks sheet loaded: {len(stocks)} rows")
                        if not stocks.empty:
                            print(f"  Columns: {list(stocks.columns)}")
                    last_sync = time.time()
                    
                    # Cleanup temp copy
                    if excel_to_read != TEMP_EXCEL and os.path.exists(temp_copy):
                        try:
                            os.remove(temp_copy)
                        except:
                            pass
                    
                    return True
                    
                except Exception as e:
                    # Cleanup temp copy on error
                    if excel_to_read != TEMP_EXCEL and os.path.exists(temp_copy):
                        try:
                            os.remove(temp_copy)
                        except:
                            pass
                    raise
                    
            except Exception as e:
                error_msg = str(e).lower()
                if "permission" in error_msg:
                    print(f"⚠️  File locked (attempt {attempt + 1}): {e}")
                elif "zip" in error_msg:
                    print(f"⚠️  File corrupted/incomplete (attempt {attempt + 1}): {e}")
                else:
                    print(f"⚠️  Load attempt {attempt + 1} failed: {e}")
                
                if attempt < max_retries - 1:
                    current_delay = retry_delays[attempt]
                    print(f"   Retrying in {current_delay}s...")
                else:
                    print(f"❌ Load failed (after {max_retries} attempts): {e}")
                    return False
# --------------------------------------------------
# POLLING THREAD
# --------------------------------------------------
last_poll_time = time.time()
poll_skip_until = time.time()

def poll_sharepoint():
    global last_poll_time, poll_skip_until
    while True:
        # Skip polling if recently asked to skip (e.g., after save_all)
        if time.time() < poll_skip_until:
            print(f"[*] Skipping poll (within skip window)")
            time.sleep(5)
            continue
        
        load_data()
        last_poll_time = time.time()
        time.sleep(30)

def skip_polling_for(seconds):
    """Tell polling thread to skip for N seconds"""
    global poll_skip_until
    poll_skip_until = time.time() + seconds
    print(f"[*] Polling will be skipped for {seconds}s")

threading.Thread(target=poll_sharepoint, daemon=True).start()
load_data()

# --------------------------------------------------
# ROUTES
# --------------------------------------------------
@app.route("/api/dashboard")
def dashboard():
    products = db.get("Products", pd.DataFrame())
    machines = db.get("Machines", pd.DataFrame())
    stock = db.get("Stock", pd.DataFrame())

    # SAFE cost map (handles "-", empty, text)
    cost_map = {}
    for _, r in products.iterrows():
        pid = str(r.get("PRODUCT_ID", "")).strip()
        if not pid or pid.lower() == "nan":
            continue

        raw_cost = r.get("PO")
        try:
            cost = float(raw_cost)
        except (TypeError, ValueError):
            cost = 0.0

        cost_map[pid] = cost

    total_units = 0
    total_value = 0

    for _, r in stock.iterrows():
        qty = float(r.get("Current_Stock", 0) or 0)
        total_units += qty
        total_value += qty * cost_map.get(str(r.get("Product_ID")), 0)

    # Process OUR_PO to propagate common fields to all rows (like /api/our-pos does)
    our_po = db.get("OUR_PO", pd.DataFrame())
    our_pos_enriched = []
    
    if not our_po.empty:
        # Helper to check if PO_ID is valid
        def is_valid_po_id(val):
            if pd.isna(val):
                return False
            val_str = str(val).strip()
            return val_str and val_str not in ("nan", "None", "")
        
        # First pass: collect common fields for each PO
        po_common_fields = {}
        current_po_id = None
        
        for _, row in our_po.iterrows():
            po_id_val = row.get("PO_ID")
            if is_valid_po_id(po_id_val):
                current_po_id = str(po_id_val).strip()
                po_common_fields[current_po_id] = {
                    "PO_ID": current_po_id,
                    "Vendor_ID": str(row.get("Vendor_ID", "")).strip() if pd.notna(row.get("Vendor_ID")) else "",
                    "Created_Date": str(row.get("Created_Date", "")).strip() if pd.notna(row.get("Created_Date")) else "",
                    "Total_Amount": float(row.get("Total_Amount", 0) or 0) if pd.notna(row.get("Total_Amount")) else 0,
                    "Status": str(row.get("Status", "Pending")).strip() or "Pending"
                }
        
        # Second pass: build enriched rows
        current_po_id = None
        for _, row in our_po.iterrows():
            po_id_val = row.get("PO_ID")
            if is_valid_po_id(po_id_val):
                current_po_id = str(po_id_val).strip()
            
            if not current_po_id:
                continue
            
            common = po_common_fields.get(current_po_id, {})
            our_pos_enriched.append({
                "PO_ID": common.get("PO_ID", ""),
                "Vendor_ID": common.get("Vendor_ID", ""),
                "Created_Date": common.get("Created_Date", ""),
                "Total_Amount": common.get("Total_Amount", 0),
                "Product_ID": str(row.get("Product_ID", "")).strip() if pd.notna(row.get("Product_ID")) else "",
                "Product_Name": str(row.get("Product_Name", "")).strip() if pd.notna(row.get("Product_Name")) else "",
                "No_of_Cases": int(row.get("No_of_Cases", 0) or 0) if pd.notna(row.get("No_of_Cases")) else 0,
                "Units_Per_Case": int(row.get("Units_Per_Case", 1) or 1) if pd.notna(row.get("Units_Per_Case")) else 1,
                "PO_Price": float(row.get("PO_Price", 0) or 0) if pd.notna(row.get("PO_Price")) else 0,
                "Line_Total": float(row.get("Line_Total", 0) or 0) if pd.notna(row.get("Line_Total")) else 0,
                "Status": common.get("Status", "Pending")
            })

    # Debug logging
    stocks_df = db.get("Stocks", pd.DataFrame())
    print(f"DEBUG Dashboard: Stocks sheet has {len(stocks_df)} rows")
    if not stocks_df.empty:
        print(f"  Columns: {list(stocks_df.columns)}")
        print(f"  Column dtypes:\n{stocks_df.dtypes}")
        print(f"  First 3 rows:")
        for idx, row in stocks_df.head(3).iterrows():
            print(f"    Row {idx}: {row.to_dict()}")
    else:
        print(f"  WARNING: Stocks DataFrame is empty!")
    
    return jsonify({
        "products": df_to_safe_dict(products),
        "machines": df_to_safe_dict(machines),
        "stock": df_to_safe_dict(stock),
        "sales": df_to_safe_dict(db.get("Sales", pd.DataFrame())),
        "purchases": df_to_safe_dict(db.get("Purchases", pd.DataFrame())),
        "refills": df_to_safe_dict(db.get("Refills", pd.DataFrame())),
        "vendors": df_to_safe_dict(db.get("Vendors", pd.DataFrame())),
        "warehouse": df_to_safe_dict(db.get("Warehouse", pd.DataFrame())),
        "purchased_products": df_to_safe_dict(db.get("Purchased_Products", pd.DataFrame())),
        "stocks": df_to_safe_dict(db.get("Stocks", pd.DataFrame())),
        "stock_assignments": df_to_safe_dict(db.get("Stock_Assignments", pd.DataFrame())),
        "our_pos": our_pos_enriched,
        "metrics": {
            "totalStockValue": round(total_value, 2),
            "totalUnits": int(total_units),
            "activeMachines": int((machines["Status"] == "Active").sum()) if not machines.empty else 0,
            "outOfStock": int((stock["Current_Stock"] <= 0).sum()) if not stock.empty else 0
        }
    })

@app.route("/api/add-product", methods=["POST"])
def add_product():
    """Add a new product to Product_Master"""
    try:
        d = request.json
        product_id = d.get("Product_ID", "").strip()
        name = d.get("Name", "").strip()
        
        if not product_id or not name:
            return jsonify(error="Product ID and Name are required"), 400
        
        products = db.get("Products", pd.DataFrame())
        
        # Check for duplicate Product_ID
        if not products.empty and "PRODUCT_ID" in products.columns:
            if product_id in products["PRODUCT_ID"].astype(str).values:
                return jsonify(error=f"Product ID '{product_id}' already exists"), 400
        
        # Create new product row matching Excel column names
        new_product = pd.DataFrame([{
            "PRODUCT_ID": product_id,
            "PRODUCT_NAME": name,
            "CATEGORY": d.get("Category", "Others"),
            "PO": float(d.get("Unit_Cost", 0)),
            "GST": float(d.get("GST", 0)),
            "MRP": float(d.get("MRP", 0)),
            "QUANTITY": int(d.get("Quantity", 0))
        }])
        
        db["Products"] = pd.concat([products, new_product], ignore_index=True)
        save_all()
        
        return jsonify(success=True, message=f"Product '{name}' added successfully")
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/update-product", methods=["POST"])
def update_product():
    """Update an existing product in Product_Master"""
    try:
        d = request.json
        product_id = d.get("Product_ID", "").strip()
        
        if not product_id:
            return jsonify(error="Product ID is required"), 400
        
        products = db.get("Products", pd.DataFrame())
        
        if products.empty or "PRODUCT_ID" not in products.columns:
            return jsonify(error="No products found"), 404
        
        mask = products["PRODUCT_ID"].astype(str) == product_id
        if not mask.any():
            return jsonify(error=f"Product '{product_id}' not found"), 404
        
        # Update the product
        if "Name" in d:
            products.loc[mask, "PRODUCT_NAME"] = d["Name"]
        if "Category" in d:
            products.loc[mask, "CATEGORY"] = d["Category"]
        if "Unit_Cost" in d:
            products.loc[mask, "PO"] = float(d["Unit_Cost"])
        if "GST" in d:
            products.loc[mask, "GST"] = float(d["GST"])
        if "MRP" in d:
            products.loc[mask, "MRP"] = float(d["MRP"])
        if "Quantity" in d:
            products.loc[mask, "QUANTITY"] = int(d["Quantity"])
        
        db["Products"] = products
        save_all()
        
        return jsonify(success=True, message=f"Product '{product_id}' updated successfully")
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/delete-product", methods=["POST"])
def delete_product():
    """Delete a product from Product_Master"""
    try:
        d = request.json
        product_id = d.get("Product_ID", "").strip()
        
        if not product_id:
            return jsonify(error="Product ID is required"), 400
        
        products = db.get("Products", pd.DataFrame())
        
        if products.empty or "PRODUCT_ID" not in products.columns:
            return jsonify(error="No products found"), 404
        
        mask = products["PRODUCT_ID"].astype(str) == product_id
        if not mask.any():
            return jsonify(error=f"Product '{product_id}' not found"), 404
        
        db["Products"] = products[~mask].reset_index(drop=True)
        save_all()
        
        return jsonify(success=True, message=f"Product '{product_id}' deleted successfully")
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/create-po", methods=["POST"])
def create_po():
    """Create a new purchase order"""
    try:
        d = request.json
        product_id = str(d.get("product_id", "")).strip()
        vendor_id = str(d.get("vendor_id", "")).strip()
        qty = int(d.get("qty", 0))
        units_per_case = int(d.get("units_per_case", 1))
        po_price = float(d.get("po_price", 0))
        notes = d.get("notes", "")
        
        if not product_id or not vendor_id:
            return jsonify(error="Product ID and Vendor ID are required"), 400
        
        purchases = db.get("Purchases", pd.DataFrame())
        
        # Generate PO number
        po_count = len(purchases) + 1 if not purchases.empty else 1
        po_number = f"PO-{datetime.now().strftime('%Y%m')}-{po_count:03d}"
        
        # Create new PO row matching Excel column names for Vendor_Purchase
        new_po = pd.DataFrame([{
            "PO Bill": po_number,
            "Date": datetime.now().strftime("%Y-%m-%d"),
            "Vendor_ID": vendor_id,
            "Product_ID": product_id,
            "Qty": qty,
            "Units_Per_Case": units_per_case,
            "Total_Units": qty * units_per_case,
            "Actual PO price": po_price,
            "Payment Status ": "Pending",
            "Notes": notes
        }])
        
        db["Purchases"] = pd.concat([purchases, new_po], ignore_index=True)
        save_all()
        
        return jsonify(success=True, message=f"PO '{po_number}' created successfully", po_number=po_number)
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/create-multi-po", methods=["POST"])
def create_multi_po():
    """
    Create purchase orders for multiple products, grouped by vendor.
    Each vendor gets a separate PO_ID. All entries stored in OUR_PO sheet.
    Structure: Common fields (PO_ID, Vendor_ID, Created_Date, Total_Amount, Status) only in first row,
    subsequent rows for the same PO have those fields empty (like Vendor_Purchase sheet).
    """
    try:
        d = request.json
        items = d.get("items", [])  # List of products with vendor_id, product_id, etc.
        
        if not items:
            return jsonify(error="No items provided"), 400
        
        # Group items by vendor_id
        vendor_groups = {}
        for item in items:
            vendor_id = str(item.get("vendor_id", "")).strip()
            if not vendor_id:
                return jsonify(error="All items must have a vendor_id"), 400
            if vendor_id not in vendor_groups:
                vendor_groups[vendor_id] = []
            vendor_groups[vendor_id].append(item)
        
        # Load or initialize OUR_PO sheet
        our_po = db.get("OUR_PO", pd.DataFrame())
        if our_po.empty:
            our_po = pd.DataFrame(columns=[
                "PO_ID", "Vendor_ID", "Created_Date", "Total_Amount", "Product_ID", "Product_Name", 
                "No_of_Cases", "Units_Per_Case", "PO_Price", "Line_Total", "Status"
            ])
        
        # Get current datetime for PO_ID generation (format: VP-YYYYMMDD-HHMMSS-VENDOR_ID)
        created_po_ids = []
        new_rows = []
        # Create a PO for each vendor
        for vendor_id, vendor_items in vendor_groups.items():
            # Generate PO_ID: VP-YYYYMMDD-HHMMSS-VENDOR_ID
            # Use a unique timestamp for each vendor group
            now = datetime.now()
            po_id = f"VP-{now.strftime('%Y%m%d-%H%M%S')}-{vendor_id}"
            created_date = now.strftime("%Y-%m-%d %H:%M:%S")
            created_po_ids.append(po_id)
            # Sleep for 1 second to ensure unique time for each PO_ID if multiple vendors
            if len(vendor_groups) > 1:
                import time
                time.sleep(1)
            # Calculate total amount for this PO: sum(Line_Total * PO_Price) for all products
            total_amount = 0.0
            for item in vendor_items:
                cases = int(item.get("no_of_cases", 0) or 0)
                units = int(item.get("units_per_case", 1) or 1)
                price_per_unit = float(item.get("po_price", 0) or 0)
                line_total = cases * units  # Total units for this product
                total_amount += line_total * price_per_unit  # Total cost for this product
            # Create a row for each product in this PO
            for idx, item in enumerate(vendor_items):
                product_id = str(item.get("product_id", "")).strip()
                product_name = str(item.get("product_name", "")).strip()
                no_of_cases = int(item.get("no_of_cases", 0) or 0)
                units_per_case = int(item.get("units_per_case", 1) or 1)
                po_price = float(item.get("po_price", 0) or 0)
                line_total = no_of_cases * units_per_case  # Total units for this line
                # Common fields only in first row, subsequent rows have None for empty cells
                # Using None ensures Excel writes empty cells, not empty strings or 'nan'
                is_first_row = idx == 0
                new_rows.append({
                    "PO_ID": po_id if is_first_row else None,
                    "Vendor_ID": vendor_id if is_first_row else None,
                    "Created_Date": created_date if is_first_row else None,
                    "Total_Amount": total_amount if is_first_row else None,
                    "Product_ID": product_id,
                    "Product_Name": product_name,
                    "No_of_Cases": no_of_cases,
                    "Units_Per_Case": units_per_case,
                    "PO_Price": po_price,
                    "Line_Total": line_total,
                    "Status": "Pending" if is_first_row else None
                })
        
        # Append new rows to OUR_PO
        if new_rows:
            new_df = pd.DataFrame(new_rows)
            # Define the exact column order we want in Excel
            column_order = [
                "PO_ID", "Vendor_ID", "Created_Date", "Total_Amount", "Product_ID", "Product_Name", 
                "No_of_Cases", "Units_Per_Case", "PO_Price", "Line_Total", "Status"
            ]
            combined = pd.concat([our_po, new_df], ignore_index=True)
            # Ensure column order and add any missing columns (add as None, not empty string)
            for col in column_order:
                if col not in combined.columns:
                    combined[col] = None
            db["OUR_PO"] = combined[column_order]
            save_all()
        
        return jsonify(
            success=True, 
            message=f"Created {len(created_po_ids)} PO(s) successfully",
            po_ids=created_po_ids,
            total_items=len(new_rows)
        )
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/our-pos", methods=["GET"])
def get_our_pos():
    """Get all POs from OUR_PO sheet with Status column"""
    try:
        our_po = db.get("OUR_PO", pd.DataFrame())
        
        if our_po.empty:
            return jsonify(pos=[])
        
        # Ensure Status column exists
        if "Status" not in our_po.columns:
            our_po["Status"] = "Pending"
        
        # Helper to check if a PO_ID value is valid (not None, NaN, empty, "nan", "None")
        def is_valid_po_id(val):
            if pd.isna(val):
                return False
            val_str = str(val).strip()
            return val_str and val_str not in ("nan", "None", "")
        
        # First pass: track common fields for each PO (from first row)
        po_common_fields = {}
        current_po_id = None
        
        for _, row in our_po.iterrows():
            po_id_val = row.get("PO_ID")
            
            # If this row has a valid PO_ID, it's the first row of a PO
            if is_valid_po_id(po_id_val):
                current_po_id = str(po_id_val).strip()
                po_common_fields[current_po_id] = {
                    "PO_ID": current_po_id,
                    "Vendor_ID": str(row.get("Vendor_ID", "")).strip() if pd.notna(row.get("Vendor_ID")) else "",
                    "Created_Date": str(row.get("Created_Date", "")).strip() if pd.notna(row.get("Created_Date")) else "",
                    "Total_Amount": float(row.get("Total_Amount", 0) or 0) if pd.notna(row.get("Total_Amount")) else 0,
                    "Status": str(row.get("Status", "Pending")).strip() or "Pending"
                }
        
        # Second pass: build enriched rows with propagated common fields
        enriched_pos = []
        current_po_id = None
        
        for _, row in our_po.iterrows():
            po_id_val = row.get("PO_ID")
            
            # Update current PO ID if this row has one
            if is_valid_po_id(po_id_val):
                current_po_id = str(po_id_val).strip()
            
            if not current_po_id:
                continue
            
            # Get common fields for this PO
            common = po_common_fields.get(current_po_id, {})
            
            enriched_pos.append({
                "PO_ID": common.get("PO_ID", ""),
                "Vendor_ID": common.get("Vendor_ID", ""),
                "Created_Date": common.get("Created_Date", ""),
                "Total_Amount": common.get("Total_Amount", 0),
                "Product_ID": str(row.get("Product_ID", "")).strip() if pd.notna(row.get("Product_ID")) else "",
                "Product_Name": str(row.get("Product_Name", "")).strip() if pd.notna(row.get("Product_Name")) else "",
                "No_of_Cases": int(row.get("No_of_Cases", 0) or 0) if pd.notna(row.get("No_of_Cases")) else 0,
                "Units_Per_Case": int(row.get("Units_Per_Case", 1) or 1) if pd.notna(row.get("Units_Per_Case")) else 1,
                "PO_Price": float(row.get("PO_Price", 0) or 0) if pd.notna(row.get("PO_Price")) else 0,
                "Line_Total": float(row.get("Line_Total", 0) or 0) if pd.notna(row.get("Line_Total")) else 0,
                "Status": common.get("Status", "Pending")
            })
        
        return jsonify(pos=enriched_pos)
    except Exception as e:
        import traceback
        print(f"ERROR in get_our_pos: {traceback.format_exc()}")
        return jsonify(error=str(e)), 500


@app.route("/api/vendor-purchases-grouped", methods=["GET"])
def get_vendor_purchases_grouped():
    """
    Get vendor purchases grouped by PO_ID.
    Structure: PO_ID -> DATE -> PURCHASE_DATE -> VENDOR_ID -> [Products] -> BATCH -> PAYMENT -> GST
    """
    try:
        purchases = db.get("Purchases", pd.DataFrame())
        
        if purchases.empty:
            return jsonify(grouped_purchases=[])
        
        # Group by PO ID
        grouped = {}
        for _, row in purchases.iterrows():
            po_id = str(row.get("PO ID", "")).strip()
            if not po_id:
                continue
            
            if po_id not in grouped:
                grouped[po_id] = {
                    "PO_ID": po_id,
                    "Date": row.get("DATE"),
                    "Purchase_Date": row.get("PURCHASE DATE"),
                    "Vendor_ID": str(row.get("VENDOR ID", "")).strip(),
                    "Batch": row.get("BATCH"),
                    "Payment_Mode": row.get("PAYMENT MODE"),
                    "Payment_Status": row.get("PAYMENT STATUS"),
                    "GST_Filed": row.get("GST FILED"),
                    "Items": [],
                    "Total_Cases": 0,
                    "Total_Quantity": 0,
                    "Total_Value": 0
                }
            
            # Add product as line item
            cases = int(row.get("CASE COUNT") or 0) if pd.notna(row.get("CASE COUNT")) else 0
            qty = int(row.get("QUANTITY") or 0) if pd.notna(row.get("QUANTITY")) else 0
            po_price = float(row.get("PO PRICE") or 0) if pd.notna(row.get("PO PRICE")) else 0
            
            grouped[po_id]["Items"].append({
                "Product_ID": str(row.get("PRODUCT ID", "")).strip(),
                "Product_Name": row.get("PRODUCT NAME") or "",
                "Units_Per_Case": int(row.get("UNIT/CASE") or 1) if pd.notna(row.get("UNIT/CASE")) else 1,
                "Case_Count": cases,
                "Quantity": qty,
                "MRP": float(row.get("MRP") or 0) if pd.notna(row.get("MRP")) else 0,
                "PO_Price": po_price
            })
            grouped[po_id]["Total_Cases"] += cases
            grouped[po_id]["Total_Quantity"] += qty
            grouped[po_id]["Total_Value"] += (po_price * cases)
        
        # Convert to list and sort by date descending
        result = list(grouped.values())
        result.sort(key=lambda x: x.get("Date") or "", reverse=True)
        
        return jsonify(grouped_purchases=result)
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/vendor-purchases", methods=["GET"])
def get_vendor_purchases_flat():
    """
    Get all vendor purchases as a flat list (all rows from Vendor_Purchase sheet)
    """
    try:
        purchases = db.get("Purchases", pd.DataFrame())
        
        if purchases.empty:
            return jsonify(purchases=[])
        
        def safe_date(val):
            """Convert date value safely, returning None for NaT/NaN"""
            if pd.isna(val):
                return None
            try:
                if hasattr(val, 'isoformat'):
                    return val.isoformat()
                return str(val)
            except:
                return None
        
        result = []
        for _, row in purchases.iterrows():
            # Skip completely empty rows
            if pd.isna(row.get("VENDOR ID")) and pd.isna(row.get("PRODUCT NAME")):
                continue
                
            result.append({
                "PO_ID": str(row.get("PO ID", "")).strip() if pd.notna(row.get("PO ID")) else "",
                "Date": safe_date(row.get("DATE")),
                "Purchase_Date": safe_date(row.get("PURCHASE DATE")),
                "Vendor_ID": str(row.get("VENDOR ID", "")).strip() if pd.notna(row.get("VENDOR ID")) else "",
                "Product_ID": str(row.get("PRODUCT ID", "")).strip() if pd.notna(row.get("PRODUCT ID")) else "",
                "Product_Name": str(row.get("PRODUCT NAME", "")).strip() if pd.notna(row.get("PRODUCT NAME")) else "",
                "Batch": str(row.get("BATCH", "")).strip() if pd.notna(row.get("BATCH")) else "",
                "Units_Per_Case": int(row.get("UNIT/CASE") or 1) if pd.notna(row.get("UNIT/CASE")) else 1,
                "Case_Count": int(row.get("CASE COUNT") or 0) if pd.notna(row.get("CASE COUNT")) else 0,
                "Quantity": int(row.get("QUANTITY") or 0) if pd.notna(row.get("QUANTITY")) else 0,
                "MRP": float(row.get("MRP") or 0) if pd.notna(row.get("MRP")) else 0,
                "PO_Price": float(row.get("PO PRICE") or 0) if pd.notna(row.get("PO PRICE")) else 0,
                "Payment_Mode": str(row.get("PAYMENT MODE", "")).strip() if pd.notna(row.get("PAYMENT MODE")) else "",
                "Payment_Status": str(row.get("PAYMENT STATUS", "")).strip() if pd.notna(row.get("PAYMENT STATUS")) else "",
                "GST_Filed": str(row.get("GST FILED", "")).strip() if pd.notna(row.get("GST FILED")) else ""
            })
        
        return jsonify(purchases=result, count=len(result))
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/generate-vendor-po-ids", methods=["POST"])
def generate_vendor_po_ids():
    """
    Generate PO_IDs for all vendor purchase records that don't have one.
    Groups by DATE + VENDOR_ID - same date & vendor get same PO_ID.
    """
    try:
        purchases = db.get("Purchases", pd.DataFrame())
        
        if purchases.empty:
            return jsonify(message="No purchases found", updated=0)
        
        # Track which PO_IDs we've already generated
        generated_ids = {}
        updated_count = 0
        
        for idx, row in purchases.iterrows():
            # Skip if already has a PO_ID
            existing_po = str(row.get("PO ID", "")).strip()
            if existing_po and existing_po.lower() not in ['nan', 'none', '']:
                continue
            
            # Get date and vendor for grouping
            date_val = row.get("DATE")
            vendor_id = str(row.get("VENDOR ID", "")).strip()
            
            if pd.isna(date_val):
                continue
            
            # Create group key (date + vendor)
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime("%Y%m%d")
            else:
                try:
                    dt = pd.to_datetime(date_val)
                    date_str = dt.strftime("%Y%m%d")
                except:
                    continue
            
            group_key = f"{date_str}_{vendor_id}"
            
            # Generate PO_ID for this group if not already done
            if group_key not in generated_ids:
                # Find the next sequence number for this date
                prefix = f"VP-{date_str}"
                existing = purchases["PO ID"].astype(str).str.strip()
                matching = existing[existing.str.startswith(prefix)]
                
                max_seq = 0
                for po_id in matching:
                    try:
                        seq = int(po_id.split("-")[-1])
                        max_seq = max(max_seq, seq)
                    except:
                        pass
                
                # Also check already generated IDs
                for gen_id in generated_ids.values():
                    if gen_id.startswith(prefix):
                        try:
                            seq = int(gen_id.split("-")[-1])
                            max_seq = max(max_seq, seq)
                        except:
                            pass
                
                generated_ids[group_key] = f"{prefix}-{str(max_seq + 1).zfill(3)}"
            
            # Update the row
            purchases.at[idx, "PO ID"] = generated_ids[group_key]
            updated_count += 1
        
        # Save back to Excel if any updates
        if updated_count > 0:
            db["Purchases"] = purchases
            save_all()
        
        return jsonify(
            message=f"Generated PO_IDs for {updated_count} records",
            updated=updated_count,
            generated_ids=list(set(generated_ids.values()))
        )
    except Exception as e:
        return jsonify(error=str(e)), 500


def generate_vendor_po_id(date_str=None):
    """Generate a unique PO ID for vendor purchases: VP-YYYYMMDD-HHMMSS-VendorID"""
    now = datetime.now()
    # If vendor_id is available, append it; else, just use timestamp
    vendor_id = None
    # Try to get vendor_id from request context if possible
    try:
        vendor_id = request.json.get("vendor_id", None)
    except:
        pass
    po_id = f"VP-{now.strftime('%Y%m%d-%H%M%S')}"
    if vendor_id:
        po_id = f"{po_id}-{vendor_id}"
    return po_id


@app.route("/api/record-vendor-purchase", methods=["POST"])
def record_vendor_purchase():
    """
    Record a complete vendor purchase with auto-generated PO_ID.
    Accepts multiple products in a single PO.
    """
    try:
        d = request.json
        
        # Common PO fields
        date = d.get("date") or datetime.now().strftime("%Y-%m-%d")
        purchase_date = d.get("purchase_date") or date
        vendor_id = str(d.get("vendor_id", "")).strip()
        batch = str(d.get("batch", "")).strip() or None
        payment_mode = str(d.get("payment_mode", "")).strip() or None
        payment_status = str(d.get("payment_status", "Pending")).strip()
        gst_filed = str(d.get("gst_filed", "No")).strip()
        
        # Line items (products)
        items = d.get("items", [])
        
        if not vendor_id:
            return jsonify(error="Vendor ID is required"), 400
        if not items or len(items) == 0:
            return jsonify(error="At least one product item is required"), 400
        
        # Generate PO ID
        po_id = generate_vendor_po_id(date)
        
        purchases = db.get("Purchases", pd.DataFrame())
        if purchases.empty:
            purchases = pd.DataFrame(columns=[
                "PO ID", "DATE", "PURCHASE DATE", "VENDOR ID", "PRODUCT ID",
                "PRODUCT NAME", "BATCH", "UNIT/CASE", "CASE COUNT", "QUANTITY",
                "MRP", "PO PRICE", "PAYMENT MODE", "PAYMENT STATUS", "GST FILED"
            ])
        
        warehouse = db.get("Warehouse", pd.DataFrame())
        if warehouse.empty:
            warehouse = pd.DataFrame(columns=["Product_ID", "Product_Name", "Available_Units", "Units_Per_Case", "Last_Received_Date", "Notes"])
        
        new_rows = []
        for idx, item in enumerate(items):
            product_id = str(item.get("product_id", "")).strip()
            product_name = str(item.get("product_name", "")).strip()
            units_per_case = int(item.get("units_per_case", 1))
            case_count = int(item.get("case_count", 0))
            quantity = case_count * units_per_case
            mrp = float(item.get("mrp", 0))
            po_price = float(item.get("po_price", 0))
            
            if not product_id or case_count <= 0:
                continue
            
            # Common PO fields only for first row (PO ID kept in all rows for status matching)
            is_first_row = len(new_rows) == 0
            
            new_rows.append({
                "PO ID": po_id,
                "DATE": date if is_first_row else "",
                "PURCHASE DATE": purchase_date if is_first_row else "",
                "VENDOR ID": vendor_id if is_first_row else "",
                "PRODUCT ID": product_id,
                "PRODUCT NAME": product_name,
                "BATCH": batch if is_first_row else "",
                "UNIT/CASE": units_per_case,
                "CASE COUNT": case_count,
                "QUANTITY": quantity,
                "MRP": mrp,
                "PO PRICE": po_price,
                "PAYMENT MODE": payment_mode if is_first_row else "",
                "PAYMENT STATUS": payment_status if is_first_row else "",
                "GST FILED": gst_filed if is_first_row else ""
            })
            
            # Update warehouse
            wh_mask = warehouse["Product_ID"].astype(str) == product_id
            if wh_mask.any():
                warehouse.loc[wh_mask, "Available_Units"] = warehouse.loc[wh_mask, "Available_Units"].fillna(0).astype(int) + quantity
                warehouse.loc[wh_mask, "Units_Per_Case"] = units_per_case
                warehouse.loc[wh_mask, "Last_Received_Date"] = date
            else:
                new_wh = pd.DataFrame([{
                    "Product_ID": product_id,
                    "Product_Name": product_name,
                    "Available_Units": quantity,
                    "Units_Per_Case": units_per_case,
                    "Last_Received_Date": date,
                    "Notes": f"From VP: {po_id}"
                }])
                warehouse = pd.concat([warehouse, new_wh], ignore_index=True)
        
        if not new_rows:
            return jsonify(error="No valid items to add"), 400
        
        db["Purchases"] = pd.concat([purchases, pd.DataFrame(new_rows)], ignore_index=True)
        db["Warehouse"] = warehouse
        save_all()
        
        total_cases = sum(r["CASE COUNT"] for r in new_rows)
        total_qty = sum(r["QUANTITY"] for r in new_rows)
        
        return jsonify(
            success=True,
            po_id=po_id,
            message=f"Created {po_id} with {len(new_rows)} products ({total_cases} cases, {total_qty} units)",
            items_count=len(new_rows),
            total_cases=total_cases,
            total_quantity=total_qty
        )
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/po-items/<po_id>", methods=["GET"])
def get_po_items(po_id):
    """Get all items/products for a specific PO from OUR_PO sheet.
    Since common fields (PO_ID, Vendor_ID, etc.) are only in the first row,
    we need to track rows that belong to the same PO by sequence.
    """
    try:
        our_po = db.get("OUR_PO", pd.DataFrame())
        
        if our_po.empty:
            return jsonify(success=True, items=[], po_date=None, vendor_id=None)
        
        # Find all rows belonging to this PO
        # First row has PO_ID, subsequent rows for same PO have None/empty PO_ID
        po_items = []
        po_date = None
        vendor_id = None
        found_po = False
        
        for _, row in our_po.iterrows():
            row_po_id = row.get("PO_ID")
            row_po_id_str = str(row_po_id).strip() if pd.notna(row_po_id) else ""
            
            # Check if this row starts a new PO
            if row_po_id_str and row_po_id_str != "nan" and row_po_id_str != "None":
                if row_po_id_str == po_id:
                    # Found the PO we're looking for
                    found_po = True
                    po_date = row.get("Created_Date", "")
                    vendor_id = str(row.get("Vendor_ID", "")).strip() if pd.notna(row.get("Vendor_ID")) else ""
                elif found_po:
                    # This is a new PO, stop collecting items
                    break
            
            # If we found our PO, collect this item
            if found_po:
                po_items.append({
                    "Product_ID": str(row.get("Product_ID", "")).strip() if pd.notna(row.get("Product_ID")) else "",
                    "Product_Name": str(row.get("Product_Name", "")).strip() if pd.notna(row.get("Product_Name")) else "",
                    "No_of_Cases": int(row.get("No_of_Cases", 0)) if pd.notna(row.get("No_of_Cases")) else 0,
                    "Units_Per_Case": int(row.get("Units_Per_Case", 1)) if pd.notna(row.get("Units_Per_Case")) else 1,
                    "PO_Price": float(row.get("PO_Price", 0)) if pd.notna(row.get("PO_Price")) else 0
                })
        
        # Format date properly
        if pd.notna(po_date):
            if hasattr(po_date, 'strftime'):
                po_date = po_date.strftime("%Y-%m-%d")
            else:
                po_date = str(po_date)[:10] if po_date else None
        
        return jsonify(success=True, items=po_items, po_date=po_date, vendor_id=vendor_id)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route("/api/record-delivery", methods=["POST"])
def record_delivery():
    """
    Record vendor delivery against a PO. Saves ALL products to Purchased_Products.
    Supports both PO products and custom products from vendor.
    """
    try:
        d = request.json
        po_id = str(d.get("po_id", "")).strip()
        vendor_id = str(d.get("vendor_id", "")).strip()
        po_date = str(d.get("po_date", "")).strip()
        purchase_date = str(d.get("purchase_date", "")).strip()
        po_products = d.get("po_products", [])  # Products from PO
        custom_products = d.get("custom_products", [])  # Custom products added by admin
        
        if not po_id:
            return jsonify(error="PO ID is required"), 400
        if not purchase_date:
            return jsonify(error="Purchase Date is required"), 400
        if not po_products and not custom_products:
            return jsonify(error="At least one product must be provided"), 400
        
        # Parse dates
        try:
            po_date_parsed = pd.to_datetime(po_date).strftime("%Y-%m-%d") if po_date else datetime.now().strftime("%Y-%m-%d")
        except:
            po_date_parsed = datetime.now().strftime("%Y-%m-%d")
        
        try:
            purchase_date_parsed = pd.to_datetime(purchase_date).strftime("%Y-%m-%d")
        except:
            return jsonify(error="Invalid Purchase Date format"), 400
        
        purchases = db.get("Purchases", pd.DataFrame())
        if purchases.empty:
            purchases = pd.DataFrame(columns=[
                "PO ID", "DATE", "PURCHASE DATE", "VENDOR ID", "PRODUCT ID",
                "PRODUCT NAME", "BATCH", "UNIT/CASE", "CASE COUNT", "QUANTITY",
                "MRP", "PO PRICE", "PAYMENT MODE", "PAYMENT STATUS", "GST FILED"
            ])
        
        purchased_products = db.get("Purchased_Products", pd.DataFrame())
        if purchased_products.empty:
            purchased_products = pd.DataFrame(columns=[
                "PO_ID", "EXP_Id", "Product_ID", "Product_Name", "Available_Units", "Units_Per_Case", 
                "Batch", "Received_Date", "Notes"
            ])
        
        new_deliveries = []
        new_purchased_items = []
        po_products_recorded = 0
        custom_products_recorded = 0
        po_products_skipped = 0
        total_units_added = 0
        
        # Process PO Products
        for product in po_products:
            case_count = int(product.get("case_count", 0))
            
            # Skip products with 0 case count
            if case_count <= 0:
                po_products_skipped += 1
                continue
            
            product_id = str(product.get("product_id", "")).strip()
            product_name = str(product.get("product_name", "")).strip()
            
            if not product_id or not product_name:
                po_products_skipped += 1
                continue
            
            batch = str(product.get("batch", "")).strip() or ""
            units_per_case = int(product.get("units_per_case", 1))
            quantity = case_count * units_per_case
            mrp = float(product.get("mrp", 0))
            po_price = float(product.get("po_price", 0))
            payment_mode = str(product.get("payment_mode", "")).strip()
            payment_status = str(product.get("payment_status", "")).strip()
            gst_filed = str(product.get("gst_filed", "")).strip()
            
            is_first_row = len(new_deliveries) == 0
            
            new_deliveries.append({
                "PO ID": po_id if is_first_row else "",
                "DATE": po_date_parsed if is_first_row else "",
                "PURCHASE DATE": purchase_date_parsed if is_first_row else "",
                "VENDOR ID": vendor_id if is_first_row else "",
                "PRODUCT ID": product_id,
                "PRODUCT NAME": product_name,
                "BATCH": batch if is_first_row else "",
                "UNIT/CASE": units_per_case,
                "CASE COUNT": case_count,
                "QUANTITY": quantity,
                "MRP": mrp,
                "PO PRICE": po_price,
                "PAYMENT MODE": payment_mode if is_first_row else "",
                "PAYMENT STATUS": payment_status if is_first_row else "",
                "GST FILED": gst_filed if is_first_row else ""
            })
            
            # Generate EXP_Id as PO_ID + Product_ID
            exp_id = f"{po_id}_{product_id}"
            
            new_purchased_items.append({
                "PO_ID": po_id if is_first_row else "",
                "EXP_Id": exp_id,
                "Product_ID": product_id,
                "Product_Name": product_name,
                "Available_Units": quantity,
                "Units_Per_Case": units_per_case,
                "Batch": batch,
                "Received_Date": purchase_date_parsed,
                "Notes": "From PO"
            })
            
            po_products_recorded += 1
            total_units_added += quantity
        
        # Process Custom Products (not in original PO)
        for product in custom_products:
            quantity = int(product.get("quantity", 0))
            
            if quantity <= 0:
                continue
            
            product_id = str(product.get("product_id", "")).strip()
            product_name = str(product.get("product_name", "")).strip()
            
            if not product_id or not product_name:
                continue
            
            batch = str(product.get("batch", "")).strip() or ""
            units_per_case = int(product.get("units_per_case", 1))
            mrp = float(product.get("mrp", 0))
            po_price = float(product.get("po_price", 0))
            
            is_first_row = len(new_deliveries) == 0
            
            # Add custom products to Purchases sheet as well
            new_deliveries.append({
                "PO ID": po_id if is_first_row else "",
                "DATE": po_date_parsed if is_first_row else "",
                "PURCHASE DATE": purchase_date_parsed if is_first_row else "",
                "VENDOR ID": vendor_id if is_first_row else "",
                "PRODUCT ID": product_id,
                "PRODUCT NAME": product_name,
                "BATCH": batch if is_first_row else "",
                "UNIT/CASE": units_per_case,
                "CASE COUNT": 0,
                "QUANTITY": quantity,
                "MRP": mrp,
                "PO PRICE": po_price,
                "PAYMENT MODE": "",
                "PAYMENT STATUS": "",
                "GST FILED": ""
            })
            
            # Generate EXP_Id as PO_ID + Product_ID
            exp_id = f"{po_id}_{product_id}"
            
            # Also add to purchased_products
            new_purchased_items.append({
                "PO_ID": po_id if is_first_row else "",
                "EXP_Id": exp_id,
                "Product_ID": product_id,
                "Product_Name": product_name,
                "Available_Units": quantity,
                "Units_Per_Case": units_per_case,
                "Batch": batch,
                "Received_Date": purchase_date_parsed,
                "Notes": "Custom Product (not in PO)"
            })
            
            custom_products_recorded += 1
            total_units_added += quantity
        
        # Save all data
        if new_deliveries:
            new_df = pd.DataFrame(new_deliveries)
            db["Purchases"] = pd.concat([purchases, new_df], ignore_index=True)
        
        if new_purchased_items:
            # Add all new items as separate records to track expiry by batch using EXP_Id
            new_purchased_df = pd.DataFrame(new_purchased_items)
            purchased_products = pd.concat([
                purchased_products,
                new_purchased_df
            ], ignore_index=True)
            
            db["Purchased_Products"] = purchased_products
        
        # Update OUR_PO Status
        our_po = db.get("OUR_PO", pd.DataFrame())
        if not our_po.empty and "Status" not in our_po.columns:
            our_po["Status"] = "Pending"
        
        if not our_po.empty and new_deliveries and new_purchased_items:
            po_first_row_idx = None
            po_product_rows = []
            current_po = None
            
            for idx, row in our_po.iterrows():
                row_po_id = row.get("PO_ID")
                row_po_id_str = str(row_po_id).strip() if pd.notna(row_po_id) else ""
                is_valid_po_id = row_po_id_str and row_po_id_str not in ("nan", "None", "")
                
                if is_valid_po_id:
                    current_po = row_po_id_str
                    if current_po == po_id:
                        po_first_row_idx = idx
                        po_product_rows = [idx]
                elif current_po == po_id:
                    po_product_rows.append(idx)
            
            # Check if data was successfully inserted in both sheets
            purchases_sheet = db.get("Purchases", pd.DataFrame())
            purchased_products_sheet = db.get("Purchased_Products", pd.DataFrame())
            
            # Verify Purchases sheet has the new entries
            purchases_has_po = False
            if not purchases_sheet.empty:
                mask = purchases_sheet["PO ID"].astype(str).str.strip() == po_id
                if mask.any():
                    purchases_has_po = True
            
            # Verify Purchased_Products sheet has the new entries
            purchased_products_has_items = False
            if not purchased_products_sheet.empty and new_purchased_items:
                for item in new_purchased_items:
                    product_id = item.get("Product_ID", "").strip()
                    mask = purchased_products_sheet["Product_ID"].astype(str).str.strip() == product_id
                    if mask.any():
                        purchased_products_has_items = True
                        break
            
            # Determine if all PO products were fully delivered
            all_delivered = True
            if po_products:
                for delivery in po_products:
                    product_id = str(delivery.get("product_id", "")).strip()
                    delivered_cases = int(delivery.get("case_count", 0))
                    
                    # Find matching product in OUR_PO
                    found = False
                    for row_idx in po_product_rows:
                        if str(our_po.at[row_idx, "Product_ID"]).strip() == product_id:
                            ordered_cases = int(our_po.at[row_idx, "No_of_Cases"] or 0)
                            if delivered_cases < ordered_cases:
                                all_delivered = False
                            found = True
                            break
                    
                    # If product not found in OUR_PO, skip it
                    if not found:
                        all_delivered = False
            
            # Update status: "Completed" if both sheets have data with either PO or custom products recorded
            if po_first_row_idx is not None:
                if purchases_has_po and purchased_products_has_items and (po_products_recorded > 0 or custom_products_recorded > 0):
                    our_po.at[po_first_row_idx, "Status"] = "Completed"
                else:
                    our_po.at[po_first_row_idx, "Status"] = "Pending"
            
            db["OUR_PO"] = our_po
        
        save_all()
        
        # Determine if both sheets were successfully updated
        both_sheets_updated = (new_deliveries and new_purchased_items and (po_products_recorded > 0 or custom_products_recorded > 0))
        
        return jsonify(
            success=True,
            message=f"Recorded delivery for PO {po_id}",
            po_products_recorded=po_products_recorded,
            custom_products_recorded=custom_products_recorded,
            po_products_skipped=po_products_skipped,
            total_units=total_units_added,
            purchased_products_updated=both_sheets_updated,
            both_sheets_updated=both_sheets_updated
        )
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/po-deliveries/<po_id>", methods=["GET"])
def get_po_deliveries(po_id):
    """Get all deliveries for a specific PO"""
    try:
        purchases = db.get("Purchases", pd.DataFrame())
        
        if purchases.empty:
            return jsonify(deliveries=[])
        
        # Filter deliveries for this PO
        mask = purchases["PO ID"].astype(str).str.strip() == po_id
        po_deliveries = purchases[mask]
        
        return jsonify(deliveries=df_to_safe_dict(po_deliveries))
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/sell", methods=["POST"])
def sell():
    d = request.json
    mid, pid, qty = d["machineId"], d["productId"], int(d.get("qty", 1))

    stock = db["Stock"]
    mask = (stock["Machine_ID"] == mid) & (stock["Product_ID"] == pid)

    if mask.any() and stock.loc[mask, "Current_Stock"].iloc[0] >= qty:
        stock.loc[mask, "Current_Stock"] -= qty

        db["Sales"] = pd.concat([db["Sales"], pd.DataFrame([{
            "Date": datetime.now().strftime("%Y-%m-%d"),
            "Machine_ID": mid,
            "Product_ID": pid,
            "Qty Sold": qty,
            "Selling_Price": d.get("price", 0)
        }])])

        save_all()
        return jsonify(success=True)

    return jsonify(error="Insufficient stock"), 400

@app.route("/api/refill", methods=["POST"])
def refill():
    d = request.json
    mid, pid, qty = d["machineId"], d["productId"], int(d["qty"])

    stock = db["Stock"]
    mask = (stock["Machine_ID"] == mid) & (stock["Product_ID"] == pid)

    # Optional metadata: Batch and Stock_ID
    batch = d.get('batch')
    stock_id = d.get('stockId') or d.get('Stock_ID')

    if mask.any():
        stock.loc[mask, "Current_Stock"] += qty
        # If existing row, optionally set Stock_ID/Batch if provided
        if stock_id is not None:
            stock.loc[mask, 'Stock_ID'] = stock_id
        if batch is not None:
            stock.loc[mask, 'Batch'] = batch
    else:
        new_row = {
            "Machine_ID": mid,
            "Product_ID": pid,
            "Current_Stock": qty
        }
        if stock_id is not None:
            new_row['Stock_ID'] = stock_id
        if batch is not None:
            new_row['Batch'] = batch
        db["Stock"] = pd.concat([stock, pd.DataFrame([new_row])])

    refill_entry = {
        "Date": datetime.now().strftime("%Y-%m-%d"),
        "Refiller_ID": d.get("refillerId", "R001"),
        "Machine_ID": mid,
        "Product_ID": pid,
        "Qty": qty
    }
    if batch is not None:
        refill_entry['Batch'] = batch
    if stock_id is not None:
        refill_entry['Stock_ID'] = stock_id

    db["Refills"] = pd.concat([db["Refills"], pd.DataFrame([refill_entry])])

    save_all()
    return jsonify(success=True)

@app.route("/api/update-stock", methods=["POST"])
def update_stock():
    """Update stock directly without creating a refill log entry"""
    d = request.json
    mid, pid, qty = d["machineId"], d["productId"], int(d["qty"])

    stock = db["Stock"]
    mask = (stock["Machine_ID"] == mid) & (stock["Product_ID"] == pid)

    if mask.any():
        stock.loc[mask, "Current_Stock"] += qty
    else:
        db["Stock"] = pd.concat([stock, pd.DataFrame([{
            "Machine_ID": mid,
            "Product_ID": pid,
            "Current_Stock": qty
        }])])

    save_all()
    return jsonify(success=True)

# --------------------------------------------------
# PURCHASED_PRODUCTS ROUTES (Received goods awaiting warehouse approval)
# --------------------------------------------------
@app.route("/api/purchased-products/items", methods=["GET"])
def get_purchased_products():
    """Get all purchased products awaiting warehouse approval - sorted by received date (earliest first for FIFO)"""
    try:
        purchased_products = db.get("Purchased_Products", pd.DataFrame())
        
        if not purchased_products.empty:
            # Sort by Received_Date (earliest first) for FIFO picking
            purchased_products['Received_Date'] = pd.to_datetime(purchased_products['Received_Date'], errors='coerce')
            purchased_products = purchased_products.sort_values('Received_Date')
        
        return jsonify(items=df_to_safe_dict(purchased_products))
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/purchased-products/move-to-warehouse", methods=["POST"])
def move_purchased_products_to_warehouse():
    """Move items from purchased products to warehouse (admin action)"""
    try:
        d = request.json
        purchased_items = d.get("items", [])  # List of {product_id, available_units, units_per_case, batch}
        
        if not purchased_items:
            return jsonify(error="No items provided"), 400
        
        purchased_products = db.get("Purchased_Products", pd.DataFrame())
        warehouse = db.get("Warehouse", pd.DataFrame())
        
        if purchased_products.empty:
            return jsonify(error="Purchased products is empty"), 400
        
        if warehouse.empty:
            warehouse = pd.DataFrame(columns=["Product_ID", "Product_Name", "Available_Units", "Units_Per_Case", "Last_Received_Date", "Notes"])
        
        moved_count = 0
        total_units = 0
        
        for item in purchased_items:
            product_id = str(item.get("product_id", "")).strip()
            units_to_move = int(item.get("available_units", 0))
            
            if not product_id or units_to_move <= 0:
                continue
            
            # Find the purchased products item
            pp_mask = purchased_products["Product_ID"].astype(str) == product_id
            if not pp_mask.any():
                continue
            
            pp_row = purchased_products[pp_mask].iloc[0]
            product_name = pp_row.get("Product_Name", "")
            units_per_case = int(pp_row.get("Units_Per_Case", 1))
            batch = pp_row.get("Batch", "")
            po_id = pp_row.get("PO_ID", "")
            
            # Add to warehouse
            wh_mask = warehouse["Product_ID"].astype(str) == product_id
            if wh_mask.any():
                warehouse.loc[wh_mask, "Available_Units"] = warehouse.loc[wh_mask, "Available_Units"].fillna(0).astype(int) + units_to_move
                warehouse.loc[wh_mask, "Units_Per_Case"] = units_per_case
                warehouse.loc[wh_mask, "Last_Received_Date"] = datetime.now().strftime("%Y-%m-%d")
            else:
                new_wh = pd.DataFrame([{
                    "Product_ID": product_id,
                    "Product_Name": product_name,
                    "Available_Units": units_to_move,
                    "Units_Per_Case": units_per_case,
                    "Last_Received_Date": datetime.now().strftime("%Y-%m-%d"),
                    "Notes": f"From Purchased Products - Batch: {batch}, PO: {po_id}"
                }])
                warehouse = pd.concat([warehouse, new_wh], ignore_index=True)
            
            # Remove from purchased products
            purchased_products = purchased_products[~pp_mask].reset_index(drop=True)
            moved_count += 1
            total_units += units_to_move
        
        db["Purchased_Products"] = purchased_products
        db["Warehouse"] = warehouse
        save_all()
        
        return jsonify(success=True, message=f"Moved {moved_count} items to warehouse", total_units=total_units)
    except Exception as e:
        return jsonify(error=str(e)), 500

# --------------------------------------------------
# STOCKS MANAGEMENT ROUTES
# --------------------------------------------------
@app.route("/api/stocks/create", methods=["POST"])
def create_stock():
    """Create a new stock (collection of products from inventory) assigned to a specific machine"""
    try:
        import uuid
        d = request.json
        stock_name = str(d.get("stock_name", "")).strip()
        machine_id = str(d.get("machine_id", "")).strip()  # REQUIRED: Machine assignment
        products = d.get("products", [])  # List of {product_id, units}
        
        if not stock_name or not products:
            return jsonify(error="Stock name and products are required"), 400
        
        if not machine_id:
            return jsonify(error="Machine assignment is required"), 400
        
        stocks = db.get("Stocks", pd.DataFrame())
        if stocks.empty:
            stocks = pd.DataFrame(columns=["Stock_ID", "Stock_Name", "Machine_ID", "Created_Date", "Status", "Total_Units", "Products_Count"])
        
        inventory = db.get("Inventory", pd.DataFrame())
        
        # Validate and prepare products
        stock_products = []
        total_units = 0
        
        for product in products:
            product_id = str(product.get("product_id", "")).strip()
            units = int(product.get("units", 0))
            
            if not product_id or units <= 0:
                continue
            
            # Check inventory availability
            inv_mask = inventory["Product_ID"].astype(str) == product_id
            if not inv_mask.any():
                return jsonify(error=f"Product {product_id} not found in inventory"), 404
            
            inv_item = inventory[inv_mask].iloc[0]
            available = int(inv_item.get("Available_Units", 0))
            
            if available < units:
                return jsonify(error=f"Insufficient stock for {product_id}. Available: {available}, Requested: {units}"), 400
            
            stock_products.append({
                "product_id": product_id,
                "product_name": inv_item.get("Product_Name", ""),
                "units": units
            })
            total_units += units
        
        if not stock_products:
            return jsonify(error="No valid products to add"), 400
        
        # Generate unique Stock_ID using timestamp (format: STK-YYYYMMDD-HHMMSS)
        stock_id = f"STK-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        
        # Create stock entry with Machine_ID
        new_stock = {
            "Stock_ID": stock_id,
            "Stock_Name": stock_name,
            "Machine_ID": machine_id,
            "Created_Date": datetime.now().strftime("%Y-%m-%d"),
            "Status": "Active",
            "Total_Units": total_units,
            "Products_Count": len(stock_products)
        }
        
        stocks = pd.concat([stocks, pd.DataFrame([new_stock])], ignore_index=True)
        db["Stocks"] = stocks
        
        # Store stock product details
        stock_assignments = db.get("Stock_Assignments", pd.DataFrame())
        if stock_assignments.empty:
            stock_assignments = pd.DataFrame(columns=["Stock_ID", "Product_ID", "Product_Name", "Units", "Assignment_Status", "Machine_ID", "Assigned_Date"])
        
        for product in stock_products:
            assignment = {
                "Stock_ID": stock_id,
                "Product_ID": product["product_id"],
                "Product_Name": product["product_name"],
                "Units": product["units"],
                "Assignment_Status": "Assigned",
                "Machine_ID": machine_id,
                "Assigned_Date": datetime.now().strftime("%Y-%m-%d")
            }
            stock_assignments = pd.concat([stock_assignments, pd.DataFrame([assignment])], ignore_index=True)
            
            # Deduct from inventory
            inv_mask = inventory["Product_ID"].astype(str) == product["product_id"]
            inventory.loc[inv_mask, "Available_Units"] = inventory.loc[inv_mask, "Available_Units"].fillna(0).astype(int) - product["units"]
            # Remove if stock becomes 0
            inventory = inventory[inventory["Available_Units"] > 0].reset_index(drop=True)
        
        db["Inventory"] = inventory
        db["Stock_Assignments"] = stock_assignments
        save_all()
        
        return jsonify(success=True, stock_id=stock_id, message=f"Created stock {stock_id} with {total_units} units")
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/stocks/list", methods=["GET"])
def list_stocks():
    """Get all stocks"""
    try:
        stocks = db.get("Stocks", pd.DataFrame())
        stock_assignments = db.get("Stock_Assignments", pd.DataFrame())
        
        stocks_list = []
        for _, stock in stocks.iterrows():
            stock_id = stock.get("Stock_ID", "")
            assignments = stock_assignments[stock_assignments["Stock_ID"] == stock_id] if not stock_assignments.empty else pd.DataFrame()
            
            stocks_list.append({
                "Stock_ID": stock_id,
                "Stock_Name": stock.get("Stock_Name", ""),
                "Machine_ID": stock.get("Machine_ID", ""),
                "Created_Date": stock.get("Created_Date", ""),
                "Status": stock.get("Status", ""),
                "Total_Units": int(stock.get("Total_Units", 0)),
                "Products_Count": int(stock.get("Products_Count", 0)),
                "Products": df_to_safe_dict(assignments)
            })
        
        return jsonify(stocks=stocks_list)
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/stocks/create-from-warehouse", methods=["POST"])
def create_stock_from_warehouse():
    """Create a new stock from warehouse items and assign to a machine"""
    try:
        d = request.json
        stock_name = str(d.get("stock_name", "")).strip()
        machine_id = str(d.get("machine_id", "")).strip()
        products = d.get("products", [])  # List of {product_id, units}
        
        if not stock_name or not machine_id:
            return jsonify(error="Stock name and machine assignment are required"), 400
        
        if not products:
            return jsonify(error="At least one product is required"), 400
        
        warehouse = db.get("Warehouse", pd.DataFrame())
        stocks = db.get("Stocks", pd.DataFrame())
        if stocks.empty:
            stocks = pd.DataFrame(columns=["Stock_ID", "Stock_Name", "Machine_ID", "Created_Date", "Status", "Total_Units", "Products_Count"])
        
        # Validate and prepare products from warehouse
        stock_products = []
        total_units = 0
        
        for product in products:
            product_id = str(product.get("product_id", "")).strip()
            units = int(product.get("units", 0))
            
            if not product_id or units <= 0:
                continue
            
            # Check warehouse availability
            if warehouse.empty:
                return jsonify(error=f"Warehouse is empty"), 400
            
            mask = warehouse["Product_ID"].astype(str) == product_id
            if not mask.any():
                return jsonify(error=f"Product {product_id} not found in warehouse"), 404
            
            warehouse_item = warehouse[mask].iloc[0]
            available = int(warehouse_item.get("Available_Units", 0))
            
            if available < units:
                return jsonify(error=f"Insufficient warehouse stock for {product_id}. Available: {available}, Requested: {units}"), 400
            
            stock_products.append({
                "product_id": product_id,
                "product_name": warehouse_item.get("Product_Name", ""),
                "units": units
            })
            total_units += units
        
        if not stock_products:
            return jsonify(error="No valid products to create stock"), 400
        
        # Generate unique Stock_ID using timestamp (format: STK-YYYYMMDD-HHMMSS)
        stock_id = f"STK-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        
        # Create stock entry with Machine_ID
        new_stock = {
            "Stock_ID": stock_id,
            "Stock_Name": stock_name,
            "Machine_ID": machine_id,
            "Created_Date": datetime.now().strftime("%Y-%m-%d"),
            "Status": "Active",
            "Total_Units": total_units,
            "Products_Count": len(stock_products)
        }
        
        stocks = pd.concat([stocks, pd.DataFrame([new_stock])], ignore_index=True)
        db["Stocks"] = stocks
        
        # Store stock product assignments
        stock_assignments = db.get("Stock_Assignments", pd.DataFrame())
        if stock_assignments.empty:
            stock_assignments = pd.DataFrame(columns=["Stock_ID", "Product_ID", "Product_Name", "Units", "Assignment_Status", "Machine_ID", "Assigned_Date"])
        
        for product in stock_products:
            assignment = {
                "Stock_ID": stock_id,
                "Product_ID": product["product_id"],
                "Product_Name": product["product_name"],
                "Units": product["units"],
                "Assignment_Status": "Assigned",
                "Machine_ID": machine_id,
                "Assigned_Date": datetime.now().strftime("%Y-%m-%d")
            }
            stock_assignments = pd.concat([stock_assignments, pd.DataFrame([assignment])], ignore_index=True)
            
            # Deduct from warehouse
            mask = warehouse["Product_ID"].astype(str) == product["product_id"]
            warehouse.loc[mask, "Available_Units"] = warehouse.loc[mask, "Available_Units"].fillna(0).astype(int) - product["units"]
            # Remove if stock becomes 0
            warehouse = warehouse[warehouse["Available_Units"] > 0].reset_index(drop=True)
        
        db["Warehouse"] = warehouse
        db["Stock_Assignments"] = stock_assignments
        save_all()
        
        return jsonify(success=True, stock_id=stock_id, message=f"Created stock {stock_id} with {total_units} units assigned to {machine_id}")
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/stocks/assign-to-machine", methods=["POST"])
def assign_stock_to_machine():
    """Assign a stock to a machine"""
    try:
        d = request.json
        stock_id = str(d.get("stock_id", "")).strip()
        machine_id = str(d.get("machine_id", "")).strip()
        
        if not stock_id or not machine_id:
            return jsonify(error="Stock ID and Machine ID are required"), 400
        
        stocks = db.get("Stocks", pd.DataFrame())
        stock_assignments = db.get("Stock_Assignments", pd.DataFrame())
        stock = db.get("Stock", pd.DataFrame())
        
        if stocks.empty or (stocks["Stock_ID"] != stock_id).all():
            return jsonify(error="Stock not found"), 404
        
        # Get all products in this stock
        assignments = stock_assignments[stock_assignments["Stock_ID"] == stock_id] if not stock_assignments.empty else pd.DataFrame()
        
        if assignments.empty:
            return jsonify(error="No products in this stock"), 400
        
        # Update stock assignments status
        for _, assignment in assignments.iterrows():
            product_id = assignment.get("Product_ID", "")
            units = int(assignment.get("Units", 0))
            
            mask = (stock_assignments["Stock_ID"] == stock_id) & (stock_assignments["Product_ID"] == product_id)
            stock_assignments.loc[mask, "Assignment_Status"] = "Assigned"
            stock_assignments.loc[mask, "Machine_ID"] = machine_id
            stock_assignments.loc[mask, "Assigned_Date"] = datetime.now().strftime("%Y-%m-%d")
            
            # Update machine stock
            s_mask = (stock["Machine_ID"] == machine_id) & (stock["Product_ID"] == product_id)
            if s_mask.any():
                stock.loc[s_mask, "Current_Stock"] = stock.loc[s_mask, "Current_Stock"].fillna(0).astype(int) + units
            else:
                new_stock_entry = pd.DataFrame([{
                    "Machine_ID": machine_id,
                    "Product_ID": product_id,
                    "Current_Stock": units
                }])
                stock = pd.concat([stock, new_stock_entry], ignore_index=True)
        
        # Update stock status
        stock_mask = stocks["Stock_ID"] == stock_id
        stocks.loc[stock_mask, "Status"] = "Assigned_to_Machine"
        
        db["Stocks"] = stocks
        db["Stock_Assignments"] = stock_assignments
        db["Stock"] = stock
        save_all()
        
        return jsonify(success=True, message=f"Stock {stock_id} assigned to machine {machine_id}")
    except Exception as e:
        return jsonify(error=str(e)), 500

# --------------------------------------------------
# WAREHOUSE ROUTES
# --------------------------------------------------
@app.route("/api/warehouse/add", methods=["POST"])
def warehouse_add():
    """Add or update warehouse stock when receiving a PO"""
    try:
        d = request.json
        product_id = str(d.get("product_id", "")).strip()
        product_name = str(d.get("product_name", "")).strip()
        units_received = int(d.get("units_received", 0))
        units_per_case = int(d.get("units_per_case", 1))
        notes = d.get("notes", "")
        
        if not product_id:
            return jsonify(error="Product ID is required"), 400
        
        warehouse = db.get("Warehouse", pd.DataFrame())
        
        if warehouse.empty:
            warehouse = pd.DataFrame(columns=["Product_ID", "Product_Name", "Available_Units", "Units_Per_Case", "Last_Received_Date", "Notes"])
        
        mask = warehouse["Product_ID"].astype(str) == product_id
        
        if mask.any():
            # Update existing row
            warehouse.loc[mask, "Available_Units"] = warehouse.loc[mask, "Available_Units"].fillna(0).astype(int) + units_received
            warehouse.loc[mask, "Units_Per_Case"] = units_per_case
            warehouse.loc[mask, "Last_Received_Date"] = datetime.now().strftime("%Y-%m-%d")
            if notes:
                warehouse.loc[mask, "Notes"] = notes
        else:
            # Add new row
            new_row = pd.DataFrame([{
                "Product_ID": product_id,
                "Product_Name": product_name,
                "Available_Units": units_received,
                "Units_Per_Case": units_per_case,
                "Last_Received_Date": datetime.now().strftime("%Y-%m-%d"),
                "Notes": notes
            }])
            warehouse = pd.concat([warehouse, new_row], ignore_index=True)
        
        db["Warehouse"] = warehouse
        save_all()
        
        return jsonify(success=True, message=f"Added {units_received} units of {product_id} to warehouse")
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/warehouse/transfer", methods=["POST"])
def warehouse_transfer():
    """Transfer units from warehouse to refiller (deduct from warehouse stock)"""
    try:
        d = request.json
        product_id = str(d.get("product_id", "")).strip()
        refiller_id = str(d.get("refiller_id", "")).strip()
        units_to_transfer = int(d.get("units", 0))
        
        if not product_id or not refiller_id:
            return jsonify(error="Product ID and Refiller ID are required"), 400
        
        warehouse = db.get("Warehouse", pd.DataFrame())
        
        if warehouse.empty:
            return jsonify(error="Warehouse is empty"), 400
        
        mask = warehouse["Product_ID"].astype(str) == product_id
        
        if not mask.any():
            return jsonify(error=f"Product {product_id} not found in warehouse"), 404
        
        available = int(warehouse.loc[mask, "Available_Units"].fillna(0).iloc[0])
        
        if available < units_to_transfer:
            return jsonify(error=f"Insufficient stock. Available: {available}, Requested: {units_to_transfer}"), 400
        
        # Deduct from warehouse
        warehouse.loc[mask, "Available_Units"] = available - units_to_transfer
        db["Warehouse"] = warehouse
        
        # Log the transfer as a refill entry (for tracking)
        db["Refills"] = pd.concat([db["Refills"], pd.DataFrame([{
            "Date": datetime.now().strftime("%Y-%m-%d"),
            "Refiller_ID": refiller_id,
            "Machine_ID": "WAREHOUSE_TRANSFER",
            "Product_ID": product_id,
            "Qty": units_to_transfer
        }])])
        
        save_all()
        
        return jsonify(success=True, message=f"Transferred {units_to_transfer} units to {refiller_id}", remaining=available - units_to_transfer)
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/warehouse/update", methods=["POST"])
def warehouse_update():
    """Update warehouse item details or adjust stock manually"""
    try:
        d = request.json
        product_id = str(d.get("product_id", "")).strip()
        
        if not product_id:
            return jsonify(error="Product ID is required"), 400
        
        warehouse = db.get("Warehouse", pd.DataFrame())
        
        if warehouse.empty:
            return jsonify(error="Warehouse is empty"), 404
        
        mask = warehouse["Product_ID"].astype(str) == product_id
        
        if not mask.any():
            return jsonify(error=f"Product {product_id} not found in warehouse"), 404
        
        # Update fields if provided
        if "available_units" in d:
            warehouse.loc[mask, "Available_Units"] = int(d["available_units"])
        if "units_per_case" in d:
            warehouse.loc[mask, "Units_Per_Case"] = int(d["units_per_case"])
        if "notes" in d:
            warehouse.loc[mask, "Notes"] = d["notes"]
        if "product_name" in d:
            warehouse.loc[mask, "Product_Name"] = d["product_name"]
        
        db["Warehouse"] = warehouse
        save_all()
        
        return jsonify(success=True, message=f"Warehouse item {product_id} updated")
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/warehouse/delete", methods=["POST"])
def warehouse_delete():
    """Remove a product from warehouse tracking"""
    try:
        d = request.json
        product_id = str(d.get("product_id", "")).strip()
        
        if not product_id:
            return jsonify(error="Product ID is required"), 400
        
        warehouse = db.get("Warehouse", pd.DataFrame())
        
        if warehouse.empty:
            return jsonify(error="Warehouse is empty"), 404
        
        mask = warehouse["Product_ID"].astype(str) == product_id
        
        if not mask.any():
            return jsonify(error=f"Product {product_id} not found in warehouse"), 404
        
        db["Warehouse"] = warehouse[~mask].reset_index(drop=True)
        save_all()
        
        return jsonify(success=True, message=f"Product {product_id} removed from warehouse")
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/warehouse/recommendation", methods=["GET"])
def warehouse_recommendation():
    """Get warehouse stock recommendations for PO planning"""
    try:
        product_id = request.args.get("product_id", "").strip()
        
        warehouse = db.get("Warehouse", pd.DataFrame())
        products = db.get("Products", pd.DataFrame())
        
        recommendations = []
        
        if product_id:
            # Single product recommendation
            if not warehouse.empty:
                mask = warehouse["Product_ID"].astype(str) == product_id
                if mask.any():
                    row = warehouse[mask].iloc[0]
                    recommendations.append({
                        "product_id": product_id,
                        "available_units": int(row.get("Available_Units") or 0),
                        "units_per_case": int(row.get("Units_Per_Case") or 1),
                        "product_name": row.get("Product_Name", "")
                    })
        else:
            # All products recommendation
            if not warehouse.empty:
                for _, row in warehouse.iterrows():
                    recommendations.append({
                        "product_id": str(row.get("Product_ID", "")),
                        "product_name": row.get("Product_Name", ""),
                        "available_units": int(row.get("Available_Units") or 0),
                        "units_per_case": int(row.get("Units_Per_Case") or 1)
                    })
        
        return jsonify(recommendations=recommendations)
    except Exception as e:
        return jsonify(error=str(e)), 500


# --------------------------------------------------
# BATCH STOCK MANAGEMENT ENDPOINTS
# --------------------------------------------------
@app.route("/api/stocks/create-batch", methods=["POST"])
def create_batch_stocks():
    """Create 4 stocks (S1, S2, S3, S4) for a batch assigned to 4 machines"""
    try:
        d = request.json
        batch_number = str(d.get("batch_number", "")).strip()
        machine_ids = d.get("machine_ids", [])  # Should be 4 machine IDs
        created_date = d.get("created_date", datetime.now().strftime("%Y-%m-%d"))
        
        if not batch_number:
            return jsonify(error="Batch number is required"), 400
        
        if len(machine_ids) != 4:
            return jsonify(error="Exactly 4 machine IDs are required for a batch"), 400
        
        # Initialize sheets if they don't exist
        stocks = db.get("Stocks", pd.DataFrame())
        if stocks.empty:
            stocks = pd.DataFrame(columns=["Stock_ID", "Batch_Number", "Stock_Name", "Machine_ID", "Created_Date", "Status", "Cover_List", "Total_Products", "Total_Units"])
        
        stock_products = db.get("Stock_Products", pd.DataFrame())
        if stock_products.empty:
            stock_products = pd.DataFrame(columns=["Stock_ID", "Batch_Number", "Stock_Name", "Cover_Name", "Product_ID", "Product_Name", "Units", "Created_Date"])
        
        created_stocks = []
        stock_names = ["S1", "S2", "S3", "S4"]
        
        for i, machine_id in enumerate(machine_ids):
            stock_name = stock_names[i]
            stock_id = f"STK-{batch_number}-{stock_name}"
            
            # Check if stock already exists
            if not stocks.empty and (stocks["Stock_ID"] == stock_id).any():
                return jsonify(error=f"Stock {stock_id} already exists"), 400
            
            new_stock = {
                "Stock_ID": stock_id,
                "Batch_Number": batch_number,
                "Stock_Name": stock_name,
                "Machine_ID": machine_id,
                "Created_Date": created_date,
                "Status": "Active",
                "Cover_List": "",  # Will be populated as covers are added
                "Total_Products": 0,
                "Total_Units": 0
            }
            
            stocks = pd.concat([stocks, pd.DataFrame([new_stock])], ignore_index=True)
            created_stocks.append(stock_id)
        
        db["Stocks"] = stocks
        db["Stock_Products"] = stock_products
        save_all()
        
        return jsonify(success=True, created_stocks=created_stocks, message=f"Created batch {batch_number} with stocks: {', '.join(created_stocks)}")
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/create-batch-full", methods=["POST"])
def create_batch_full():
    """Create complete batch with all stocks, covers, and products - matching Excel Stocks sheet structure
    
    Excel Structure:
    Batch | Date | Machine | Stock | cover | cover_status | product_id | product_name | units | Status
    
    Each product gets its own row. Batch/Date/Machine/Stock info only in first product of stock.
    """
    try:
        d = request.json
        batch_number = str(d.get("batch_number", "")).strip()
        machine_ids = d.get("machine_ids", [])  # [M1, M2, M3, M4]
        created_date = d.get("created_date", datetime.now().strftime("%Y-%m-%d"))
        stocks_data = d.get("stocks", {})  # {S1: {machine, covers: {C1: [products], C2: [...]}}}
        
        print(f"\n{'='*60}")
        print(f"DEBUG: create_batch_full called")
        print(f"batch_number={batch_number} (type: {type(batch_number).__name__})")
        print(f"machine_ids={machine_ids}")
        print(f"created_date={created_date}")
        print(f"stocks_data keys={list(stocks_data.keys())}")
        
        # Validation
        if not batch_number:
            raise Exception("Batch number is required")
        if not machine_ids or len(machine_ids) != 4:
            raise Exception(f"Exactly 4 machine IDs required, got {len(machine_ids) if machine_ids else 0}")
        if not stocks_data:
            raise Exception("At least one stock with covers and products is required")
        
        # Convert batch_number to numeric for consistency with existing data (1.0, 2.0 format)
        try:
            batch_num_float = float(batch_number)
            batch_number_for_db = batch_num_float
        except (ValueError, TypeError):
            batch_number_for_db = batch_number  # Keep as string if not numeric
        
        # Get Stocks sheet
        stocks = db.get("Stocks", pd.DataFrame())
        
        # Define required columns matching Excel structure exactly (with spaces and lowercase)
        required_columns = ["Batch", "Date", "Machine", "Stock", "cover", "cover status", "product id", "product name", "units", "Status"]
        
        if stocks.empty:
            stocks = pd.DataFrame(columns=required_columns)
        else:
            # Ensure all required columns exist (check both underscore and space versions)
            for idx, col in enumerate(required_columns):
                col_underscore = col.replace(" ", "_")
                # Check if column exists in either format
                exists = col in stocks.columns or col_underscore in stocks.columns
                
                if not exists:
                    print(f"DEBUG: Adding missing column '{col}' to Stocks")
                    stocks[col] = None
                elif col_underscore in stocks.columns and col not in stocks.columns:
                    # Rename underscore version to space version
                    print(f"DEBUG: Renaming '{col_underscore}' to '{col}'")
                    stocks.rename(columns={col_underscore: col}, inplace=True)
        
        total_rows_added = 0
        stock_names = ["S1", "S2", "S3", "S4"]
        
        # Track if this is the first row of the ENTIRE batch (not just first of each stock)
        first_row_overall = True
        # Track first row per stock (for Stock name - only appears once per stock like Machine)
        first_row_per_stock = {stock: True for stock in stock_names}
        
        print(f"DEBUG: Processing {len(stock_names)} stocks: {stock_names}")
        
        # Process each stock (S1, S2, S3, S4)
        for i, stock_name in enumerate(stock_names):
            if i >= len(machine_ids):
                print(f"DEBUG: Skipping {stock_name} - not enough machine_ids")
                break
            
            machine_id = machine_ids[i]
            stock_form_data = stocks_data.get(stock_name, {})
            
            print(f"\nDEBUG: [STOCK {i}] Processing {stock_name} with machine {machine_id}")
            print(f"DEBUG:   stock_form_data={stock_form_data}")
            
            if not isinstance(stock_form_data, dict):
                print(f"  WARNING: Invalid data format for {stock_name}, skipping")
                continue
            
            covers_dict = stock_form_data.get("covers", {})
            print(f"  Covers: {list(covers_dict.keys())}")
            
            # Process covers and products
            for cover_name, products_list in covers_dict.items():
                if not products_list:
                    continue
                
                print(f"  Processing cover {cover_name} with {len(products_list)} products")
                first_cover_row = True  # Flag to show cover info only once per cover
                
                for product in products_list:
                    product_id = str(product.get("product_id", "")).strip()
                    product_name = str(product.get("product_name", "")).strip()
                    units = product.get("units", 0)
                    
                    # Validate
                    if not product_id or not product_name:
                        print(f"    WARNING: Skipping product with missing ID or name")
                        continue
                    
                    try:
                        units = int(units) if units else 0
                    except (ValueError, TypeError):
                        units = 0
                    
                    if units <= 0:
                        print(f"    WARNING: Skipping {product_name} - units <= 0")
                        continue
                    
                    # Create row matching Excel structure:
                    # Batch only in FIRST row of ENTIRE batch
                    # Machine and Stock only in FIRST row per stock
                    # Cover info only in first row of each cover
                    new_row = {
                        "Batch": batch_number_for_db if first_row_overall else None,  # Only first row of entire batch
                        "Date": created_date if first_row_overall else None,  # Only first row of entire batch
                        "Machine": machine_id if first_row_per_stock[stock_name] else None,  # Only first row per stock
                        "Stock": stock_name if first_row_per_stock[stock_name] else None,  # Only first row per stock
                        "cover": cover_name if first_cover_row else None,
                        "cover status": "covered" if first_cover_row else None,  # Default value - with space
                        "product id": product_id,  # With space
                        "product name": product_name,  # With space
                        "units": units,
                        "Status": "Active" if first_row_overall else None  # Only first row of entire batch
                    }
                    
                    print(f"    Adding: {product_name} ({units} units)")
                    
                    stocks = pd.concat([stocks, pd.DataFrame([new_row])], ignore_index=True)
                    total_rows_added += 1
                    first_row_overall = False  # Mark first row as used (globally for entire batch)
                    first_row_per_stock[stock_name] = False  # Mark this stock's first row as used
                    first_cover_row = False
        
        print(f"\nDEBUG: Total rows added: {total_rows_added}")
        
        # DEBUG: Show what we're about to save
        print(f"DEBUG: Stocks dataframe before save:")
        print(f"  Total rows in stocks: {len(stocks)}")
        print(f"  Last 4 rows:")
        print(stocks[['Batch', 'Stock', 'product id', 'units']].tail(4))
        print(f"  Checking if batch numbers exist for new rows:")
        new_batch_rows = stocks[stocks['Batch'] == batch_number_for_db]
        print(f"    Rows with Batch {batch_number_for_db}: {len(new_batch_rows)}")
        print(f"    Stock values: {new_batch_rows['Stock'].tolist()}")
        
        # Write debug to file
        import os
        debug_file = os.path.join(os.environ.get('TEMP', 'C:\\temp'), 'batch_debug.log')
        with open(debug_file, "a") as f:
            f.write(f"\n\n{'='*60}\nBatch {batch_number} Creation Debug\n{'='*60}\n")
            f.write(f"Total rows added: {total_rows_added}\n")
            f.write(f"Total rows in stocks DF: {len(stocks)}\n")
            f.write(f"batch_number_for_db = {batch_number_for_db} (type: {type(batch_number_for_db).__name__})\n")
            f.write(f"Last 4 rows before save:\n")
            f.write(stocks[['Batch', 'Stock', 'product id', 'units']].tail(4).to_string())
            f.write(f"\n\nRows with Batch {batch_number_for_db}: {len(new_batch_rows)}\n")
            f.write(f"Stock values for new batch: {new_batch_rows['Stock'].tolist()}\n")
        
        if total_rows_added == 0:
            raise Exception("No products were added to the batch")
        
        # Save to database
        db["Stocks"] = stocks
        save_all()
        
        print(f"DEBUG: Successfully saved batch {batch_number} with {total_rows_added} rows")
        print(f"{'='*60}\n")
        
        return jsonify(
            success=True,
            batch_number=batch_number,
            total_rows_added=total_rows_added,
            message=f"Successfully created batch {batch_number} with {total_rows_added} product rows"
        )
    except Exception as e:
        print(f"ERROR in create_batch_full: {str(e)}")
        import traceback
        traceback.print_exc()
        print(f"{'='*60}\n")
        return jsonify(error=f"Error creating batch: {str(e)}"), 500


@app.route("/api/stocks/add-cover", methods=["POST"])
def add_cover_to_stock():
    """Add a cover (C1, C2, etc.) to a stock"""
    try:
        d = request.json
        stock_id = str(d.get("stock_id", "")).strip()
        cover_name = str(d.get("cover_name", "")).strip()
        
        if not stock_id or not cover_name:
            return jsonify(error="Stock ID and cover name are required"), 400
        
        stocks = db.get("Stocks", pd.DataFrame())
        
        if stocks.empty or (stocks["Stock_ID"] != stock_id).all():
            return jsonify(error="Stock not found"), 404
        
        # Find stock and update cover list
        stock_mask = stocks["Stock_ID"] == stock_id
        stock = stocks[stock_mask].iloc[0]
        covers = stock.get("Cover_List", "")
        covers_list = [c.strip() for c in covers.split(",") if c.strip()] if covers else []
        
        if cover_name in covers_list:
            return jsonify(error=f"Cover {cover_name} already exists in this stock"), 400
        
        covers_list.append(cover_name)
        new_cover_list = ", ".join(covers_list)
        
        stocks.loc[stock_mask, "Cover_List"] = new_cover_list
        db["Stocks"] = stocks
        save_all()
        
        return jsonify(success=True, message=f"Cover {cover_name} added to stock {stock_id}")
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/add-product", methods=["POST"])
def add_product_to_cover():
    """Add a product to a cover in a stock"""
    try:
        d = request.json
        stock_id = str(d.get("stock_id", "")).strip()
        cover_name = str(d.get("cover_name", "")).strip()
        product_id = str(d.get("product_id", "")).strip()
        product_name = str(d.get("product_name", "")).strip()
        units = int(d.get("units", 0))
        
        if not all([stock_id, cover_name, product_id, product_name, units > 0]):
            return jsonify(error="Stock ID, cover name, product ID, product name, and units are required"), 400
        
        stocks = db.get("Stocks", pd.DataFrame())
        stock_products = db.get("Stock_Products", pd.DataFrame())
        
        if stocks.empty or (stocks["Stock_ID"] != stock_id).all():
            return jsonify(error="Stock not found"), 404
        
        stock = stocks[stocks["Stock_ID"] == stock_id].iloc[0]
        batch_number = stock.get("Batch_Number", "")
        stock_name = stock.get("Stock_Name", "")
        created_date = stock.get("Created_Date", datetime.now().strftime("%Y-%m-%d"))
        
        # Add product entry
        new_product = {
            "Stock_ID": stock_id,
            "Batch_Number": batch_number,
            "Stock_Name": stock_name,
            "Cover_Name": cover_name,
            "Product_ID": product_id,
            "Product_Name": product_name,
            "Units": units,
            "Created_Date": created_date
        }
        
        stock_products = pd.concat([stock_products, pd.DataFrame([new_product])], ignore_index=True)
        
        # Update stock totals
        total_products = 0
        total_units = 0
        
        if not stock_products.empty:
            stock_filter = stock_products["Stock_ID"] == stock_id
            total_products = len(stock_products[stock_filter])
            total_units = int(stock_products[stock_filter]["Units"].sum())
        
        stock_mask = stocks["Stock_ID"] == stock_id
        stocks.loc[stock_mask, "Total_Products"] = total_products
        stocks.loc[stock_mask, "Total_Units"] = total_units
        
        db["Stocks"] = stocks
        db["Stock_Products"] = stock_products
        save_all()
        
        return jsonify(success=True, message=f"Product {product_id} added to cover {cover_name}")
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/get-covers/<stock_id>", methods=["GET"])
def get_stock_covers(stock_id):
    """Get all covers and products in a specific stock"""
    try:
        stock_products = db.get("Stock_Products", pd.DataFrame())
        
        if stock_products.empty:
            return jsonify(covers=[])
        
        stock_filter = stock_products["Stock_ID"] == stock_id
        stock_data = stock_products[stock_filter]
        
        if stock_data.empty:
            return jsonify(covers=[])
        
        # Group by cover
        covers = {}
        for _, row in stock_data.iterrows():
            cover = row.get("Cover_Name", "")
            if cover not in covers:
                covers[cover] = []
            
            covers[cover].append({
                "product_id": row.get("Product_ID", ""),
                "product_name": row.get("Product_Name", ""),
                "units": int(row.get("Units", 0))
            })
        
        covers_list = [{"cover_name": cover, "products": products} for cover, products in covers.items()]
        return jsonify(covers=covers_list)
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/suggest-products/<batch_number>/<cover_name>", methods=["GET"])
def suggest_products_from_batch(batch_number, cover_name):
    """Get product suggestions from a specific batch and cover (typically Batch 1, C1)"""
    try:
        stock_products = db.get("Stock_Products", pd.DataFrame())
        
        if stock_products.empty:
            return jsonify(suggestions=[])
        
        # Find products in the specified batch's first stock (S1) and cover
        filter_mask = (
            (stock_products["Batch_Number"] == batch_number) &
            (stock_products["Stock_Name"] == "S1") &
            (stock_products["Cover_Name"] == cover_name)
        )
        
        suggestions_data = stock_products[filter_mask]
        
        suggestions = []
        for _, row in suggestions_data.iterrows():
            suggestions.append({
                "product_id": row.get("Product_ID", ""),
                "product_name": row.get("Product_Name", ""),
                "available_units": int(row.get("Units", 0)),
                "batch_from": batch_number
            })
        
        return jsonify(suggestions=suggestions)
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/get-batches", methods=["GET"])
def get_all_batches():
    """Get all batch numbers for reference"""
    try:
        stocks = db.get("Stocks", pd.DataFrame())
        
        if stocks.empty:
            return jsonify(batches=[])
        
        batches = stocks["Batch_Number"].unique().tolist()
        return jsonify(batches=sorted(batches))
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/get-batch-details/<batch_number>", methods=["GET"])
def get_batch_details(batch_number):
    """Get details of all stocks in a batch"""
    try:
        stocks = db.get("Stocks", pd.DataFrame())
        stock_products = db.get("Stock_Products", pd.DataFrame())
        
        if stocks.empty:
            return jsonify(stocks=[])
        
        batch_filter = stocks["Batch_Number"] == batch_number
        batch_stocks = stocks[batch_filter]
        
        stocks_list = []
        for _, stock in batch_stocks.iterrows():
            stock_id = stock.get("Stock_ID", "")
            products_filter = stock_products["Stock_ID"] == stock_id
            products = stock_products[products_filter]
            
            stocks_list.append({
                "stock_id": stock_id,
                "stock_name": stock.get("Stock_Name", ""),
                "machine_id": stock.get("Machine_ID", ""),
                "created_date": stock.get("Created_Date", ""),
                "status": stock.get("Status", ""),
                "cover_list": stock.get("Cover_List", ""),
                "total_products": int(stock.get("Total_Products", 0)),
                "total_units": int(stock.get("Total_Units", 0)),
                "products": df_to_safe_dict(products)
            })
        
        return jsonify(stocks=stocks_list)
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/decrease-purchased-units", methods=["POST"])
def decrease_purchased_units():
    """Decrease available units from Purchased_Products and remove row if it hits zero"""
    try:
        d = request.json
        items_to_decrease = d.get("items", [])  # List of {exp_id, units_used}
        
        if not items_to_decrease:
            return jsonify(error="No items provided"), 400
        
        purchased_products = db.get("Purchased_Products", pd.DataFrame())
        
        if purchased_products.empty:
            return jsonify(success=True, message="No purchased products to update")
        
        rows_updated = 0
        
        for item in items_to_decrease:
            exp_id = str(item.get("exp_id", "")).strip()
            units_used = int(item.get("units_used", 0))
            
            if not exp_id or units_used <= 0:
                continue
            
            # Find row with matching EXP_Id
            mask = purchased_products["EXP_Id"].astype(str).str.strip() == exp_id
            
            if not mask.any():
                continue
            
            idx = mask.idxmax()
            current_units = int(purchased_products.at[idx, "Available_Units"] or 0)
            new_units = current_units - units_used
            
            if new_units <= 0:
                # Don't delete - just set units to 0
                purchased_products.at[idx, "Available_Units"] = 0
                rows_updated += 1
            else:
                # Update available units
                purchased_products.at[idx, "Available_Units"] = new_units
                rows_updated += 1
        
        db["Purchased_Products"] = purchased_products
        save_all()
        
        return jsonify(
            success=True,
            rows_updated=rows_updated,
            message=f"Updated {rows_updated} rows (no rows deleted, units set to 0 instead)"
        )
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/decrease-from-sources", methods=["POST"])
def decrease_from_sources():
    """Decrease available units from multiple sources (previous batches, warehouse, purchased products)
    
    Expected sources format:
    [
        {type: 'previous_batch', batch_number: 'B1', stock_name: 'S1', cover_name: 'C1', product_id: 'P1', units: 10},
        {type: 'warehouse', product_id: 'P1', units: 5},
        {type: 'purchased_product', exp_id: 'EXP123', product_id: 'P1', units: 3}
    ]
    """
    try:
        d = request.json
        sources = d.get("sources", [])  # List of sources to decrease from
        
        if not sources:
            return jsonify(success=True, results={'processed': [], 'failed': []}, message="No sources to process")
        
        stocks = db.get("Stocks", pd.DataFrame())
        warehouse = db.get("Warehouse_Stock", pd.DataFrame())
        purchased_products = db.get("Purchased_Products", pd.DataFrame())
        
        results = {'processed': [], 'failed': []}
        
        for source in sources:
            source_type = source.get("type", "").lower()
            units_to_decrease = int(source.get("units", 0))
            product_id = source.get("product_id", "")
            
            if not source_type or units_to_decrease <= 0 or not product_id:
                results['failed'].append({
                    'source': source,
                    'reason': 'Invalid source data: missing type, units, or product_id'
                })
                continue
            
            try:
                if source_type == 'previous_batch':
                    # Decrease from Stocks sheet for a specific batch/stock/cover
                    batch_number = source.get("batch_number", "")
                    stock_name = source.get("stock_name", "")
                    cover_name = source.get("cover_name", "")
                    
                    if not batch_number or not stock_name or not cover_name:
                        results['failed'].append({
                            'source': source,
                            'reason': 'Missing batch_number, stock_name, or cover_name'
                        })
                        continue
                    
                    # Normalize batch_number for comparison (convert to float then back to match format in DB)
                    try:
                        batch_num_normalized = float(batch_number)
                    except (ValueError, TypeError):
                        batch_num_normalized = batch_number
                    
                    # Find the row in Stocks sheet matching all criteria
                    # Stocks sheet format: Batch, Date, Machine, Stock, cover, cover status, product id, product name, units, Status
                    
                    # First, infer Batch, Stock, and Cover for rows where they might be NaN
                    # Batch: only first row of entire batch has Batch filled
                    # Stock: only first product of each stock has Stock filled  
                    # Cover: only first product of each cover has cover filled
                    stocks_inferred = stocks.copy()
                    current_batch = None
                    current_stock = None
                    current_cover = None
                    for idx in stocks_inferred.index:
                        batch_val = stocks_inferred.at[idx, "Batch"]
                        stock_val = stocks_inferred.at[idx, "Stock"]
                        cover_val = stocks_inferred.at[idx, "cover"]
                        
                        # Infer batch
                        if pd.notna(batch_val):
                            current_batch = batch_val
                        elif current_batch is not None:
                            stocks_inferred.at[idx, "Batch"] = current_batch
                        
                        # Infer stock (reset when Batch changes or when we explicitly see a new stock)
                        if pd.notna(stock_val):
                            current_stock = stock_val
                        elif current_stock is not None and pd.notna(current_batch):
                            stocks_inferred.at[idx, "Stock"] = current_stock
                        
                        # Infer cover (reset when Batch changes or when we explicitly see a new cover)
                        if pd.notna(cover_val):
                            current_cover = cover_val
                        elif current_cover is not None and pd.notna(current_batch):
                            stocks_inferred.at[idx, "cover"] = current_cover
                    
                    # Convert all batch values to float for comparison to handle "1" vs "1.0" vs 1.0
                    batch_mask = pd.to_numeric(stocks_inferred["Batch"], errors='coerce') == batch_num_normalized
                    stock_mask = stocks_inferred["Stock"].astype(str).str.strip() == stock_name
                    cover_mask = stocks_inferred["cover"].astype(str).str.strip() == cover_name
                    product_mask = stocks_inferred["product id"].astype(str).str.strip() == product_id
                    
                    # First try direct match (for rows where Stock is not NaN)
                    combined_mask = batch_mask & stock_mask & cover_mask & product_mask
                    
                    # If direct match fails, try inferring stock using Machine column (each stock assigned to specific machine)
                    if not combined_mask.any():
                        # Try matching batch, cover, and product_id only (ignoring stock)
                        fallback_mask = batch_mask & cover_mask & product_mask
                        
                        if fallback_mask.any():
                            # Build a machine-to-stock mapping for this batch
                            # Each stock is assigned to exactly one machine, so Machine column can be used to infer stock
                            batch_val_for_mapping = stocks_inferred.iloc[fallback_mask.idxmax()]["Batch"]
                            batch_rows = stocks_inferred[pd.to_numeric(stocks_inferred["Batch"], errors='coerce') == pd.to_numeric(batch_val_for_mapping, errors='coerce')]
                            
                            machine_to_stock = {}
                            for _, row in batch_rows.iterrows():
                                machine = safe_value(row.get("Machine"))
                                stock = safe_value(row.get("Stock"))
                                if machine and stock and stock in ['S1', 'S2', 'S3', 'S4']:
                                    machine_to_stock[machine] = stock
                            
                            # Now find which machine/stock this product belongs to by searching nearby rows
                            matching_indices = fallback_mask[fallback_mask].index
                            valid_idx = None
                            
                            for idx_candidate in matching_indices:
                                # Search in a window around this row to find the Machine value
                                # Since each stock section is ~15+ rows, look back to find it
                                window_start = max(0, idx_candidate - 20)
                                window_end = min(len(stocks_inferred), idx_candidate + 5)
                                
                                # Search backward from idx_candidate to find nearest Machine
                                found_machine = None
                                for search_idx in range(idx_candidate, window_start - 1, -1):
                                    machine_val = safe_value(stocks_inferred.at[search_idx, "Machine"])
                                    if machine_val and machine_val in machine_to_stock:
                                        # Make sure this machine is in the same batch
                                        batch_check = stocks_inferred.at[search_idx, "Batch"]
                                        if pd.to_numeric(batch_check, errors='coerce') == batch_num_normalized:
                                            found_machine = machine_val
                                            break
                                
                                if found_machine and found_machine in machine_to_stock:
                                    inferred_stock = machine_to_stock[found_machine]
                                    if inferred_stock == stock_name:
                                        valid_idx = idx_candidate
                                        break
                            
                            if valid_idx is not None:
                                combined_mask = pd.Series([False] * len(stocks))
                                combined_mask.iloc[valid_idx] = True
                    
                    if not combined_mask.any():
                        results['failed'].append({
                            'source': source,
                            'reason': f"Product {product_id} not found in batch {batch_number}/{stock_name}/{cover_name}"
                        })
                        continue
                    
                    # Get the index from the ORIGINAL stocks dataframe, not the inferred one
                    idx_in_original = combined_mask.idxmax()
                    idx = idx_in_original
                    current_units = int(stocks.at[idx, "units"] or 0)
                    new_units = current_units - units_to_decrease
                    
                    if new_units <= 0:
                        # Don't delete the row - just set units to 0
                        stocks.at[idx, "units"] = 0
                    else:
                        # Update units
                        stocks.at[idx, "units"] = new_units
                    
                    results['processed'].append({
                        'type': 'previous_batch',
                        'batch': batch_number,
                        'stock': stock_name,
                        'cover': cover_name,
                        'product_id': product_id,
                        'units_decreased': units_to_decrease,
                        'row_deleted': False,
                        'units_now_zero': new_units <= 0
                    })
                
                elif source_type == 'warehouse':
                    # Decrease from Warehouse_Stock sheet
                    # Find matching product in warehouse (assuming product_id is in an ID column)
                    warehouse_mask = warehouse["Product_ID"].astype(str).str.strip() == product_id if "Product_ID" in warehouse.columns else pd.Series([False] * len(warehouse))
                    
                    if not warehouse_mask.any():
                        results['failed'].append({
                            'source': source,
                            'reason': f"Product {product_id} not found in warehouse"
                        })
                        continue
                    
                    idx = warehouse_mask.idxmax()
                    current_units = int(warehouse.at[idx, "Available_Units"] or 0)
                    new_units = current_units - units_to_decrease
                    
                    if new_units <= 0:
                        # Don't delete - just set units to 0
                        warehouse.at[idx, "Available_Units"] = 0
                    else:
                        # Update available units
                        warehouse.at[idx, "Available_Units"] = new_units
                    
                    results['processed'].append({
                        'type': 'warehouse',
                        'product_id': product_id,
                        'units_decreased': units_to_decrease,
                        'row_deleted': new_units <= 0
                    })
                
                elif source_type == 'purchased_product':
                    # Decrease from Purchased_Products sheet
                    exp_id = source.get("exp_id", "")
                    
                    if not exp_id:
                        results['failed'].append({
                            'source': source,
                            'reason': 'Missing exp_id for purchased_product'
                        })
                        continue
                    
                    # Find row with matching EXP_Id
                    mask = purchased_products["EXP_Id"].astype(str).str.strip() == exp_id
                    
                    if not mask.any():
                        results['failed'].append({
                            'source': source,
                            'reason': f"EXP_Id {exp_id} not found in purchased products"
                        })
                        continue
                    
                    idx = mask.idxmax()
                    current_units = int(purchased_products.at[idx, "Available_Units"] or 0)
                    new_units = current_units - units_to_decrease
                    
                    if new_units <= 0:
                        # Don't delete - just set units to 0
                        purchased_products.at[idx, "Available_Units"] = 0
                    else:
                        # Update available units
                        purchased_products.at[idx, "Available_Units"] = new_units
                    
                    results['processed'].append({
                        'type': 'purchased_product',
                        'exp_id': exp_id,
                        'product_id': product_id,
                        'units_decreased': units_to_decrease,
                        'row_deleted': False,
                        'units_now_zero': new_units <= 0
                    })
                
                else:
                    results['failed'].append({
                        'source': source,
                        'reason': f"Unknown source type: {source_type}"
                    })
            
            except Exception as source_error:
                results['failed'].append({
                    'source': source,
                    'reason': f"Error processing source: {str(source_error)}"
                })
        
        # Reset indices after any deletions
        stocks = stocks.reset_index(drop=True)
        warehouse = warehouse.reset_index(drop=True)
        purchased_products = purchased_products.reset_index(drop=True)
        
        # After all decreases, check which batches should be marked as Inactive
        # Only mark a batch as Inactive if ALL its products have 0 units
        if not stocks.empty:
            # First, infer Batch column (only first row of batch has it filled, rest are NaN)
            stocks_with_inferred_batch = stocks.copy()
            current_batch = None
            for idx in stocks_with_inferred_batch.index:
                batch_val = stocks_with_inferred_batch.at[idx, "Batch"]
                if pd.notna(batch_val):
                    current_batch = batch_val
                elif current_batch is not None:
                    stocks_with_inferred_batch.at[idx, "Batch"] = current_batch
            
            # Get unique batch numbers (now all rows have batch inferred)
            unique_batches = pd.to_numeric(stocks_with_inferred_batch["Batch"], errors='coerce').dropna().unique()
            
            for batch_num in unique_batches:
                # Get all rows for this batch (including inferred batch values)
                batch_mask = pd.to_numeric(stocks_with_inferred_batch["Batch"], errors='coerce') == batch_num
                batch_rows = stocks_with_inferred_batch[batch_mask]
                
                if len(batch_rows) > 0:
                    # Check if ALL products in this batch have 0 units
                    batch_units = pd.to_numeric(batch_rows["units"], errors='coerce').fillna(0)
                    all_units_zero = (batch_units == 0).all()
                    
                    # Mark batch as Inactive ONLY if ALL products have 0 units
                    if all_units_zero:
                        # Find first row of this batch and set Status to "Inactive"
                        # Get the index from the ORIGINAL stocks dataframe
                        first_row_idx = batch_rows.index[0]
                        stocks.at[first_row_idx, "Status"] = "Inactive"
                    else:
                        # Ensure batch is marked as Active if any product still has units
                        first_row_idx = batch_rows.index[0]
                        stocks.at[first_row_idx, "Status"] = "Active"
        
        # Save all updates
        db["Stocks"] = stocks
        db["Warehouse_Stock"] = warehouse
        db["Purchased_Products"] = purchased_products
        save_all()
        
        # Determine success/partial success
        all_failed = len(results['failed']) > 0 and len(results['processed']) == 0
        
        if all_failed:
            return jsonify(
                success=False,
                results=results,
                message=f"Failed to process {len(results['failed'])} sources"
            ), 400
        else:
            return jsonify(
                success=True,
                results=results,
                message=f"Processed {len(results['processed'])} sources, {len(results['failed'])} failed"
            )
    
    except Exception as e:
        print(f"ERROR in decrease_from_sources: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/get-previous-patterns", methods=["GET"])
def get_previous_stock_patterns():
    """Get previously created stock patterns for copying"""
    try:
        stocks = db.get("Stocks", pd.DataFrame())
        stock_products = db.get("Stock_Products", pd.DataFrame())
        
        if stocks.empty:
            return jsonify(patterns=[])
        
        patterns = []
        
        # Group by batch and stock combination
        for _, stock in stocks.iterrows():
            stock_id = stock.get("Stock_ID", "")
            stock_name = stock.get("Stock_Name", "")
            batch_number = stock.get("Batch_Number", "")
            
            # Get all products and covers for this stock
            products_mask = stock_products["Stock_ID"] == stock_id
            stock_products_data = stock_products[products_mask]
            
            # Group by cover
            covers = {}
            for _, prod in stock_products_data.iterrows():
                cover_name = prod.get("Cover_Name", "")
                if cover_name not in covers:
                    covers[cover_name] = []
                
                covers[cover_name].append({
                    "product_id": prod.get("Product_ID", ""),
                    "product_name": prod.get("Product_Name", ""),
                    "units": int(prod.get("Units", 0))
                })
            
            patterns.append({
                "pattern_id": f"{batch_number}-{stock_name}",
                "batch_number": batch_number,
                "stock_name": stock_name,
                "label": f"Same as {stock_name} (Batch {batch_number})",
                "covers": covers,
                "total_products": int(stock.get("Total_Products", 0)),
                "total_units": int(stock.get("Total_Units", 0))
            })
        
        return jsonify(patterns=patterns)
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/get-next-cover-name/<stock_id>", methods=["GET"])
def get_next_cover_name(stock_id):
    """Get the next cover name in sequence (C, C2, C3, ..., Cn)"""
    try:
        stock_products = db.get("Stock_Products", pd.DataFrame())
        
        if stock_products.empty:
            return jsonify(next_cover="C")
        
        # Get all covers for this stock
        stock_mask = stock_products["Stock_ID"] == stock_id
        stock_covers_data = stock_products[stock_mask]
        
        if stock_covers_data.empty:
            return jsonify(next_cover="C")
        
        existing_covers = stock_covers_data["Cover_Name"].unique().tolist()
        existing_covers = [str(c).strip() for c in existing_covers if c and str(c).strip()]
        
        if not existing_covers:
            return jsonify(next_cover="C")
        
        # Generate cover names: C, C2, C3, ..., Cn
        cover_names = ["C"]
        for i in range(2, 100):  # Support up to C99
            cover_names.append(f"C{i}")
        
        # Find the first unused cover name
        for cover in cover_names:
            if cover not in existing_covers:
                return jsonify(next_cover=cover)
        
        # Fallback (shouldn't reach here)
        return jsonify(next_cover=f"C{len(existing_covers) + 1}")
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/get-batch-products/<batch_number>", methods=["GET"])
def get_batch_products(batch_number):
    """Get all products for a batch"""
    try:
        stock_products = db.get("Stock_Products", pd.DataFrame())
        
        if stock_products.empty:
            return jsonify(products=[])
        
        # Filter by batch number
        batch_mask = stock_products["Batch_Number"] == batch_number
        batch_products_data = stock_products[batch_mask]
        
        if batch_products_data.empty:
            return jsonify(products=[])
        
        products = []
        for _, prod in batch_products_data.iterrows():
            products.append({
                "stock_id": prod.get("Stock_ID", ""),
                "stock_name": prod.get("Stock_Name", ""),
                "cover_name": prod.get("Cover_Name", ""),
                "product_id": prod.get("Product_ID", ""),
                "product_name": prod.get("Product_Name", ""),
                "units": int(prod.get("Units", 0)),
                "created_date": prod.get("Created_Date", "")
            })
        
        return jsonify(products=products)
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/update-status", methods=["POST"])
def update_stock_status():
    """Update stock status (Active -> Inactive when all units are zero)"""
    try:
        d = request.json
        stock_id = str(d.get("stock_id", "")).strip()
        new_status = str(d.get("status", "")).strip()
        
        if not stock_id:
            return jsonify(error="Stock ID is required"), 400
        
        stocks = db.get("Stocks", pd.DataFrame())
        stock_products = db.get("Stock_Products", pd.DataFrame())
        
        if stocks.empty:
            return jsonify(error="Stocks sheet is empty"), 404
        
        # Find stock
        stock_mask = stocks["Stock_ID"] == stock_id
        if not stock_mask.any():
            return jsonify(error=f"Stock {stock_id} not found"), 404
        
        stock_idx = stock_mask.idxmax()
        
        # Check if all products in this stock have zero units
        stock_products_mask = stock_products["Stock_ID"] == stock_id
        stock_products_data = stock_products[stock_products_mask]
        
        total_units = 0
        if not stock_products_data.empty:
            total_units = int(stock_products_data["Units"].sum() or 0)
        
        # Auto-change to Inactive if all units are zero
        if total_units == 0:
            new_status = "Inactive"
        elif total_units > 0 and new_status == "Inactive":
            # Prevent manual change to Inactive if units still exist
            return jsonify(error="Cannot mark as Inactive while units still exist"), 400
        
        # Update status
        stocks.at[stock_idx, "Status"] = new_status
        stocks.at[stock_idx, "Total_Units"] = total_units
        
        db["Stocks"] = stocks
        save_all()
        
        return jsonify(
            success=True,
            stock_id=stock_id,
            new_status=new_status,
            total_units=total_units,
            message=f"Stock {stock_id} status updated to {new_status}"
        )
    except Exception as e:
        print(f"ERROR in update_stock_status: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error=str(e)), 500


@app.route("/api/stocks/get-batch-suggestions", methods=["GET"])
def get_batch_suggestions():
    """Get batch suggestions for all stock-cover-product combinations.
    Returns a dict keyed by 'stock-cover-productId' with warehouse_units for each.
    """
    try:
        stocks = db.get("Stocks", pd.DataFrame())
        warehouse = db.get("Warehouse", pd.DataFrame())
        
        suggestions = {}
        
        if not stocks.empty:
            for _, row in stocks.iterrows():
                stock_name = safe_value(row.get("Stock", "")) or ""
                cover_name = safe_value(row.get("cover", "")) or ""
                product_id = str(safe_value(row.get("product id", "")) or "")
                product_name = safe_value(row.get("product name", "")) or ""
                
                # Skip if no product_id
                if not product_id:
                    continue
                
                # Build key as stock-cover-productId
                key = f"{stock_name}-{cover_name}-{product_id}"
                
                # Get warehouse units for this product
                warehouse_units = 0
                if not warehouse.empty:
                    wh_mask = warehouse["Product_ID"].astype(str) == product_id
                    if wh_mask.any():
                        warehouse_units = int(warehouse.loc[wh_mask, "Available_Units"].iloc[0] or 0)
                
                suggestions[key] = {
                    "product_name": product_name,
                    "warehouse_units": warehouse_units,
                    "batches": []
                }
        
        return jsonify(suggestions=suggestions)
    except Exception as e:
        print(f"ERROR in get_batch_suggestions: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error=str(e)), 500


def safe_value(val, default=""):
    """Convert NaN and None to appropriate defaults. Returns None for null values."""
    if pd.isna(val):
        return None if default == "" else default
    val_str = str(val).strip()
    if val_str.lower() in ('nan', 'none', ''):
        return None
    return val_str if isinstance(val, str) else val


@app.route("/api/stocks/get-suggestions-detailed", methods=["POST"])
def get_suggestions_detailed():
    """Get detailed suggestions for a product in a specific stock/cover.
    Filters batches by even/odd grouping and returns only previous batches (batch_num < current_batch_num).
    Returns: {previous_batches, warehouse, purchased_products}
    """
    try:
        d = request.json
        stock_name = str(d.get("stock_name", "")).strip()
        cover_name = str(d.get("cover_name", "")).strip()
        product_id = str(d.get("product_id", "")).strip()
        current_batch_number = str(d.get("current_batch_number", "")).strip()
        
        stocks = db.get("Stocks", pd.DataFrame())
        warehouse = db.get("Warehouse", pd.DataFrame())
        purchased_products = db.get("Purchased_Products", pd.DataFrame())
        
        # Extract current batch number and determine odd/even grouping
        current_batch_num = 0
        is_current_odd = None
        if current_batch_number:
            try:
                # Handle formats like "1", "1.0", "batch_1" by extracting the number before the decimal
                import re
                match = re.search(r'(\d+)(?:\.(\d+))?', str(current_batch_number))
                if match:
                    # Get the integer part
                    current_batch_num = int(match.group(1))
                    is_current_odd = current_batch_num % 2 == 1
            except:
                current_batch_num = 0
        
        # 1. Find previous batches containing this product
        previous_batches = []
        if not stocks.empty:
            stocks_copy = stocks.copy()
            
            # First pass: Fill NaN batch numbers by forward-filling from previous valid batch
            # This handles cases where S2, S3, S4 rows don't have batch numbers but belong to a batch
            current_batch = None
            for idx in stocks_copy.index:
                batch_val = stocks_copy.at[idx, "Batch"]
                if pd.notna(batch_val):
                    current_batch = batch_val
                elif current_batch is not None and pd.isna(batch_val):
                    # Infer batch from previous row
                    stocks_copy.at[idx, "Batch"] = current_batch
            
            # Second pass: Fill NaN stock names by forward-filling within stock groups
            # This handles cases where C2, C3 rows don't have stock but belong to a stock group
            current_stock = None
            for idx in stocks_copy.index:
                stock_val = stocks_copy.at[idx, "Stock"]
                cover_val = stocks_copy.at[idx, "cover"]
                
                if pd.notna(stock_val):
                    # Store current stock
                    current_stock = stock_val
                elif pd.notna(cover_val) and current_stock is not None:
                    # If cover exists and stock is known, infer stock (but reset on blank cover rows)
                    stocks_copy.at[idx, "Stock"] = current_stock
                elif pd.isna(cover_val):
                    # Reset on blank rows
                    current_stock = None
            
            # Third pass: Fill NaN product IDs by forward-filling within same stock
            # This handles cases where C2, C3 rows don't have product_id but belong to same product
            current_product_id = None
            current_stock = None
            for idx in stocks_copy.index:
                product_id_val = stocks_copy.at[idx, "product id"]
                stock_val = stocks_copy.at[idx, "Stock"]
                
                if pd.notna(product_id_val):
                    # Store current product and stock
                    current_product_id = product_id_val
                    current_stock = stock_val
                elif pd.notna(stock_val) and stock_val == current_stock and current_product_id is not None:
                    # If same stock and product is known, infer product_id
                    stocks_copy.at[idx, "product id"] = current_product_id
                elif pd.isna(stock_val):
                    # Reset if stock changed
                    current_stock = None
                    current_product_id = None
            
            # Find all stocks containing this product
            product_mask = stocks_copy["product id"].astype(str) == product_id
            stocks_with_product = stocks_copy[product_mask]
            
            # Get unique batches containing this product from THE SAME STOCK AND SAME COVER
            unique_batches = {}
            for idx, stock_row in stocks_with_product.iterrows():
                batch_number = stock_row.get("Batch", "")
                stock_name_from_db = stock_row.get("Stock", "")
                cover_from_db = stock_row.get("cover", "")
                available_units = int(stock_row.get("units", 0) or 0)
                
                # Handle blank stock - infer from nearby C rows that come BEFORE this row
                # This handles C2, C3, etc. rows where stock might be blank/NaN
                stock_is_blank = pd.isna(stock_name_from_db) or (isinstance(stock_name_from_db, str) and stock_name_from_db.strip() == '')
                
                if stock_is_blank and cover_from_db in ['C2', 'C3', 'C4', 'C5']:
                    # Find the corresponding C (not C2, C3, etc) row for same product that comes BEFORE this row
                    # This ensures we get the RIGHT stock (S2-C before S2-C2, not S1-C)
                    try:
                        current_idx_pos = list(stocks_copy.index).index(idx)
                        # Search backwards up to 10 rows to find nearest C row
                        c_rows_before = stocks_copy.iloc[max(0, current_idx_pos - 10):current_idx_pos]
                        c_matches = c_rows_before[(c_rows_before["product id"].astype(str) == product_id) & 
                                                  (c_rows_before["cover"] == 'C')]
                        if not c_matches.empty:
                            # Use stock from the nearest C row before this row (last one in the range)
                            inferred_stock = safe_value(c_matches.iloc[-1].get("Stock"))
                            if inferred_stock:
                                stock_name_from_db = inferred_stock
                            # Also infer batch if needed
                            if pd.isna(batch_number):
                                inferred_batch = c_matches.iloc[-1].get("Batch")
                                if pd.notna(inferred_batch):
                                    batch_number = inferred_batch
                    except:
                        pass  # If index lookup fails, continue with blank stock
                
                # Skip rows with invalid batch numbers or stock names
                if pd.isna(batch_number):
                    continue
                
                batch_str = safe_value(batch_number, "Unknown") or "Unknown"
                # Ensure batch_str is always a string (not a float)
                batch_str = str(batch_str)
                stock_name_from_db = safe_value(stock_name_from_db, "") or ""
                cover_from_db = safe_value(cover_from_db, "") or ""
                
                # Skip if no valid stock or batch, or no available units
                if not stock_name_from_db or not batch_str or batch_str == "Unknown" or available_units <= 0:
                    continue
                
                # IMPORTANT: Only include batches from the SAME STOCK AND SAME COVER
                if stock_name_from_db != stock_name:
                    continue
                
                if cover_from_db != cover_name:
                    continue
                
                # Extract batch number and apply grouping filter
                try:
                    # Handle formats like "1", "1.0", "batch_1" by extracting the integer part
                    import re
                    match = re.search(r'(\d+)(?:\.(\d+))?', str(batch_str))
                    if match:
                        batch_num = int(match.group(1))
                    else:
                        batch_num = 0
                except:
                    batch_num = 0
                
                # IMPORTANT: Apply even/odd grouping filter
                # Only show batches from the same odd/even group as current batch
                if is_current_odd is not None:
                    batch_is_odd = batch_num % 2 == 1
                    if batch_is_odd != is_current_odd:
                        continue  # Skip batches from different group
                    
                    # Only show batches that come BEFORE current batch in the same group
                    if batch_num >= current_batch_num:
                        continue
                
                # Create unique key from batch + stock combination (strict cover filter applied above)
                batch_key = f"{batch_str}-{stock_name_from_db}"
                
                if batch_key not in unique_batches:
                    unique_batches[batch_key] = {
                        "batch_number": batch_str,
                        "batch_num": batch_num,
                        "stock_name": stock_name_from_db,
                        "cover_name": cover_from_db,
                        "product_id": product_id,
                        "units_available": available_units
                    }
            
            previous_batches = list(unique_batches.values())
            # Sort by batch number descending (most recent first)
            previous_batches.sort(key=lambda x: x["batch_num"], reverse=True)
        
        # 2. Get warehouse availability for this product
        warehouse_info = None
        if not warehouse.empty:
            wh_mask = warehouse["Product_ID"].astype(str) == product_id
            if wh_mask.any():
                wh_row = warehouse[wh_mask].iloc[0]
                available_units = int(wh_row.get("Available_Units", 0) or 0)
                if available_units > 0:
                    warehouse_info = {
                        "product_id": product_id,
                        "product_name": safe_value(wh_row.get("Product_Name", "")) or "",
                        "units_available": available_units,
                        "units_per_case": int(wh_row.get("Units_Per_Case", 1) or 1)
                    }
        
        # 3. Get purchased products for this product (most common first)
        purchased_items = []
        if not purchased_products.empty:
            pp_mask = purchased_products["Product_ID"].astype(str) == product_id
            pp_data = purchased_products[pp_mask]
            
            for _, pp_row in pp_data.iterrows():
                available_units = int(pp_row.get("Available_Units", 0) or 0)
                if available_units > 0:
                    batch_val = safe_value(pp_row.get("Batch", "")) or ""
                    # Ensure batch is always a string
                    batch_val = str(batch_val) if batch_val else ""
                    purchased_items.append({
                        "exp_id": safe_value(pp_row.get("EXP_Id", "")) or "",
                        "product_id": product_id,
                        "product_name": safe_value(pp_row.get("Product_Name", "")) or "",
                        "batch": batch_val,
                        "available_units": available_units,
                        "units_per_case": int(pp_row.get("Units_Per_Case", 1) or 1),
                        "received_date": safe_value(pp_row.get("Received_Date", "")) or ""
                    })
        
        return jsonify(suggestions={
            "previous_batches": previous_batches,
            "warehouse": warehouse_info,
            "purchased_products": purchased_items
        })
    except Exception as e:
        print(f"ERROR in get_suggestions_detailed: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(error=str(e)), 500


# --------------------------------------------------
# RUN SERVER
# --------------------------------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=PORT, debug=False)